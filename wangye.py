import streamlit as st
import pandas as pd
import os
import logging
import re
import streamlit.components.v1 as components
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles import numbers
import base64
import sys
from io import BytesIO
import requests
import tempfile
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
from PIL import Image
import io

# ============================
# 初始化设置
# ============================
# 设置页面配置
st.set_page_config(
    page_title="数据处理工具",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.info("启动数据处理工具。")


# ============================
# 学业桥数据处理相关工具函数
# ============================

# ======== 路径兼容函数 =========
def resource_path(relative_path):
    """兼容 PyCharm 开发环境 和 PyInstaller 打包后的路径"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


# ======== 加载学校数据 =========
try:
    school_data_path = resource_path("school_data.xlsx")
    school_df = pd.read_excel(school_data_path)
    VALID_SCHOOL_NAMES = set(school_df['学校名称'].dropna().str.strip())
    logging.info(f"成功加载 {len(VALID_SCHOOL_NAMES)} 个有效学校名称")
except Exception as e:
    logging.error(f"读取 school_data.xlsx 出错：{e}")
    VALID_SCHOOL_NAMES = set()
    st.warning("学校数据加载失败，学校名称检查功能将不可用")

# ======== 加载招生专业数据 =========
try:
    major_data_path = resource_path("招生专业.xlsx")
    major_df = pd.read_excel(major_data_path)
    VALID_MAJOR_COMBOS = set(major_df['招生专业'].dropna().astype(str).str.strip())
    logging.info(f"成功加载 {len(VALID_MAJOR_COMBOS)} 个有效专业组合")
except Exception as e:
    logging.error(f"读取 招生专业.xlsx 出错：{e}")
    VALID_MAJOR_COMBOS = set()
    st.warning("专业数据加载失败，专业匹配功能将不可用")


def check_school_name(name):
    if pd.isna(name) or not str(name).strip():
        return '学校名称为空'
    return '匹配' if name.strip() in VALID_SCHOOL_NAMES else '不匹配'


def check_major_combo(major, level):
    if pd.isna(major) or pd.isna(level):
        return "数据缺失"
    combo = f"{str(major).strip()}{str(level).strip()}"
    return "匹配" if combo in VALID_MAJOR_COMBOS else "不匹配"


def convert_selection_requirement_from_requirement(req):
    """
    依据上传文件中的报考要求转换为选科要求说明与次选科目（与 docx 规范一致）。
    1. 报考要求：不限 → 选科要求说明：不限科目专业组，次选科目：空白
    2. 报考要求仅为单个字（如"化""政"）→ 选科要求说明：单科、多科均需选考，次选科目=报考要求
    3. 报考要求中包含"且"（如"物且化"、"物且化且生"）→ 选科要求说明：单科、多科均需选考，次选科目为去掉"且"
    4. 报考要求中包含"或"（如"物或化"、"物或化或生"）→ 选科要求说明：多门选考，次选科目为去掉"或"
    """
    if pd.isna(req) or not str(req).strip():
        return "不限科目专业组", ""
    s = str(req).strip()
    if "不限" in s:
        return "不限科目专业组", ""
    if len(s) == 1:
        return "单科、多科均需选考", s
    if "且" in s:
        return "单科、多科均需选考", s.replace("且", "")
    if "或" in s:
        return "多门选考", s.replace("或", "")
    return "", ""


def _to_text(value):
    """转换为文本格式（学业桥工具用）"""
    if value is None or (value != 0 and not value):
        return ''
    text = str(value).lstrip('^').strip().lstrip("'")
    return text


def _get_first_subject(category):
    """根据科类取首选科目（学业桥工具用）"""
    if not category:
        return ''
    c = str(category)
    if '物理类' in c or '物理' in c:
        return '物'
    if '历史类' in c or '历史' in c:
        return '历'
    return ''


def _normalize_kele(kele):
    """转换招生科类：物理→物理类，历史→历史类，其他科类直接返回。"""
    if kele is None or (isinstance(kele, str) and not kele.strip()):
        return ''
    k = str(kele).strip()
    if k == '物理':
        return '物理类'
    if k == '历史':
        return '历史类'
    return k


# 专业组代码按省份转换：无专业组 / 招生代码+专业组编号 / 招生代码=专业组代码 / 招生代码+（专业组编号）
PROVINCE_NO_GROUP = {'河北', '辽宁', '山东', '浙江', '重庆', '贵州', '青海', '新疆', '西藏'}
PROVINCE_CODE_PLUS_GROUP = {'吉林'}   # 招生代码+专业组编号，如 320401、0200001
PROVINCE_CODE_EQUALS_GROUP = {'湖北', '江苏', '上海', '海南', '天津'}  # 招生代码=专业组代码，如 320401


def _convert_group_code_by_province(province, zhaosheng_code, group_no):
    """
    按省份转换专业组代码。
    1. 河北、辽宁、山东、浙江、重庆、贵州、青海、新疆、西藏：无专业组代码，无需转换，返回空
    2. 海南、吉林：招生代码+专业组编号（如 320401、0200001）
    3. 湖北、江苏、上海、天津：招生代码=专业组代码（如 320401）
    4. 其余省份：招生代码+（专业组编号）（如 3204（01）、0200（001））
    """
    p = (province or '').strip()
    code = _to_text(zhaosheng_code or '')
    group = _to_text(group_no or '')
    if p in PROVINCE_NO_GROUP:
        return ''
    if p in PROVINCE_CODE_PLUS_GROUP:
        return (code or '') + (group or '')
    if p in PROVINCE_CODE_EQUALS_GROUP:
        return code or ''
    # 其余省份：招生代码+（专业组编号）
    if not group:
        return code or ''
    return (code or '') + '（' + group + '）'


# 学业桥上传文件从第一行（标题行）开始校验，必须包含以下字段
XUEYEQIAO_UPLOAD_COLUMNS = [
    '数据类型', '年份', '省份', '批次', '科类', '院校名称', '院校原始名称', '招生代码', '专业组编号',
    '专业代码', '招生类型', '专业名称', '报考要求', '专业备注', '招生计划人数', '最低分', '最低位次',
    '最高分', '平均分', '录取人数'
]

# 学业桥导出文件第3行标题列（与上传字段映射后的导出格式）
XUEYEQIAO_EXPORT_HEADERS = [
    '学校名称', '省份', '招生专业', '专业方向（选填）', '专业备注（选填）', '一级层次', '招生科类', '招生批次',
    '招生类型（选填）', '最高分', '最低分', '平均分', '最低分位次（选填）', '招生人数（选填）', '数据来源',
    '专业组代码', '首选科目', '选科要求', '次选科目', '专业代码', '招生代码',
    '最低分数区间低', '最低分数区间高', '最低分数区间位次低', '最低分数区间位次高', '录取人数（选填）',
    '修改后备注', '备注修改说明'
]

# 学业桥导出文件第1行合并单元格备注内容（A1-U1，行高220磅）
XUEYEQIAO_EXPORT_NOTE = (
    '备注：请删除示例后再填写；\n'
    '1.省份：必须填写各省份简称，例如：北京、内蒙古，不能带有市、省、自治区、空格、特殊字符等\n'
    '2.科类：浙江、上海限定"综合、艺术类、体育类"，内蒙古限定"文科、理科、蒙授文科、蒙授理科、艺术类、艺术文、艺术理、体育类、体育文、体育理、蒙授艺术、蒙授体育"，其他省份限定"文科、理科、艺术类、艺术文、艺术理、体育类、体育文、体育理"\n'
    '3.批次：（以下为19年使用批次）\n'
    '河北、内蒙古、吉林、江苏、安徽、福建、江西、河南、湖北、广西、重庆、四川、贵州、云南、西藏、陕西、甘肃、宁夏、新疆限定本科提前批、本科一批、本科二批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；\n'
    '黑龙江、湖南、青海限定本科提前批、本科一批、本科二批、本科三批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；\n'
    '山西限定本科一批A段、本科一批B段、本科二批A段、本科二批B段、本科二批C段、专科批、国家专项计划本科批、地方专项计划本科批；\n'
    '浙江限定普通类提前批、平行录取一段、平行录取二段、平行录取三段\n'
    '4.招生人数：仅能填写数字\n'
    '5.最高分、最低分、平均分：仅能填写数字，保留小数后两位，且三者顺序不能改变，最低分为必填项，其中艺术类和体育类分数为文化课分数\n'
    '6.一级层次：限定"本科、专科（高职）"，该部分为招生专业对应的专业层次\n'
    '7.最低分位次：仅能填写数字;\n'
    '8.数据来源：必须限定——官方考试院、大红本数据、学校官网、销售、抓取、圣达信、优志愿、学业桥\n'
    '9.选科要求：不限科目专业组;多门选考;单科、多科均需选考\n'
    '10.选科科目必须是科目的简写（物、化、生、历、地、政、技）\n'
    '11.2020北京、海南，17-19上海仅限制本科专业组代码必填\n'
    '12.新八省首选科目必须选择（物理或历史）\n'
    '13.分数区间仅限北京'
)


def map_upload_row_to_export(row):
    """
    将上传文件的一行映射为导出文件格式。
    字段映射：学校名称←院校名称，招生专业←专业名称，招生科类←科类，专业组代码←专业组编号等；
    首选科目由科类经 _get_first_subject 得到；选科要求、次选科目由报考要求经 convert_selection_requirement_from_requirement 转换。
    """
    new_row = {}
    new_row['学校名称'] = row.get('院校名称', '') or ''
    new_row['省份'] = row.get('省份', '') or ''
    new_row['招生专业'] = row.get('专业名称', '') or ''
    new_row['专业方向（选填）'] = row.get('专业方向（选填）', '') or ''
    new_row['专业备注（选填）'] = row.get('专业备注', '') or ''
    new_row['一级层次'] = row.get('一级层次', '') or ''
    # 招生科类：物理→物理类，历史→历史类，其他直接转换
    kele_raw = row.get('科类', '') or ''
    new_row['招生科类'] = _normalize_kele(kele_raw)
    new_row['招生批次'] = row.get('批次', '') or ''
    new_row['招生类型（选填）'] = row.get('招生类型', '') or ''
    new_row['最高分'] = row.get('最高分', '') or ''
    new_row['最低分'] = row.get('最低分', '') or ''
    new_row['平均分'] = row.get('平均分', '') or ''
    new_row['最低分位次（选填）'] = row.get('最低位次', '') or ''
    new_row['招生人数（选填）'] = row.get('招生计划人数', '') or ''
    new_row['数据来源'] = row.get('数据来源', '') or ''
    # 专业组代码按省份转换
    province = row.get('省份', '') or ''
    zhaosheng_code = row.get('招生代码', '') or ''
    group_no = row.get('专业组编号', '') or row.get('专业组代码', '')
    new_row['专业组代码'] = _convert_group_code_by_province(province, zhaosheng_code, group_no)
    cat = row.get('科类', '') or ''
    new_row['首选科目'] = _get_first_subject(cat)
    req = row.get('报考要求', '') or ''
    sel_desc, second = convert_selection_requirement_from_requirement(req)
    new_row['选科要求'] = sel_desc
    new_row['次选科目'] = second
    new_row['专业代码'] = _to_text(row.get('专业代码', ''))
    new_row['招生代码'] = _to_text(row.get('招生代码', ''))
    new_row['最低分数区间低'] = row.get('最低分数区间低', '') or ''
    new_row['最低分数区间高'] = row.get('最低分数区间高', '') or ''
    new_row['最低分数区间位次低'] = row.get('最低分数区间位次低', '') or ''
    new_row['最低分数区间位次高'] = row.get('最低分数区间位次高', '') or ''
    new_row['录取人数（选填）'] = row.get('录取人数', '') or ''
    # 修改后备注和备注修改说明放在最后两列
    new_row['修改后备注'] = row.get('修改后备注', '') or ''
    new_row['备注修改说明'] = row.get('备注检查结果', '') or ''
    return new_row


CUSTOM_WHITELIST = {
    "宏福校区", "沙河校区", "中外合作办学", "珠海校区", "江北校区", "津南校区", "开封校区",
    "联合办学", "校企合作", "合作办学", "威海校区", "深圳校区", "苏州校区", "平果校区",
    "江南校区", "合川校区", "长安校区", "崇安校区", "南校区", "东校区", "都市园艺", "甘肃兰州"
}

TYPO_DICT = {
    "教助": "救助",
    "指辉": "指挥",
    "料学": "科学",
    "话言": "语言",
    "5十3": "5+3",
    "5十3一体化": "5+3一体化",
    "“5十3”一体化": "“5+3”一体化",
    "5+31体化": "5+3一体化",
    "5+3体化": "5+3一体化",
    "色言": "色盲",
    "NIT": "NIIT",
    "色育": "色盲",
    "人围": "入围",
    "项月": "项目",
    "币范类": "师范类",
    "投课": "授课",
    "就薄": "就读",
    "电请": "申请",
    "中国面": "中国画",
    "火数民族": "少数民族",
    "色自": "色盲",
    "色盲色弱申报": "色盲色弱慎报",
    "数学与应用数笑": "数学与应用数学",
    "法学十": "法学+",
    "浣海校区": "滨海校区",
    "中溴": "中澳"
}

REGEX_PATTERNS = {
    'excess_punct': re.compile(r'[，、。！？；,;.!? ]+'),
    'outer_punct': re.compile(r'^[，、。！？；,;.!? ]+|[，、。！？；,;.!? ]+$'),
    'consecutive_right': re.compile(r'）{2,}')
}
NESTED_PAREN_PATTERN = re.compile(r'（（(.*?)））')
CONSECUTIVE_REPEAT_PATTERN = re.compile(r'（(.+?)）\s*（\1）')


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def normalize_brackets(text):
    """统一各种括号为中文括号并处理不完整括号"""
    if pd.isna(text) or not str(text).strip():
        return text
    text = str(text).strip()

    # 替换所有括号变体为中文括号
    text = re.sub(r'[{\[【]', '（', text)  # 左括号
    text = re.sub(r'[}\]】]', '）', text)  # 右括号
    text = re.sub(r'[<《]', '（', text)  # 左书名号替换为左括号
    text = re.sub(r'[>》]', '）', text)  # 右书名号替换为右括号

    return text


def clean_outer_punctuation(text):
    """清理最外层括号外的标点符号"""
    if pd.isna(text) or not str(text).strip():
        return text
    text = str(text).strip()
    text = REGEX_PATTERNS['outer_punct'].sub('', text)
    parts = re.split(r'(（.*?）)', text)
    cleaned_parts = []
    for part in parts:
        if part.startswith('（') and part.endswith('）'):
            cleaned_parts.append(part)
        else:
            cleaned_parts.append(REGEX_PATTERNS['outer_punct'].sub('', part))
    return ''.join(cleaned_parts)


def check_score_consistency(row):
    """检查分数一致性：最高分 >= 平均分 >= 最低分"""
    issues = []
    try:
        max_score = float(row['最高分']) if pd.notna(row['最高分']) else None
        avg_score = float(row['平均分']) if pd.notna(row['平均分']) else None
        min_score = float(row['最低分']) if pd.notna(row['最低分']) else None

        if max_score is not None and avg_score is not None and max_score < avg_score:
            issues.append(f"最高分({max_score}) < 平均分({avg_score})")

        if max_score is not None and min_score is not None and max_score < min_score:
            issues.append(f"最高分({max_score}) < 最低分({min_score})")

        if avg_score is not None and min_score is not None and avg_score < min_score:
            issues.append(f"平均分({avg_score}) < 最低分({min_score})")

    except (ValueError, TypeError) as e:
        issues.append(f"分数格式错误: {str(e)}")

    return '；'.join(issues) if issues else '无问题'


def analyze_and_fix(text):
    if pd.isna(text) or not str(text).strip():
        return text, []

    text = normalize_brackets(text)
    text = clean_outer_punctuation(text)
    issues = []

    if text in CUSTOM_WHITELIST:
        return text, []

    # ========== 括号成对修正 ==========
    text_list = list(text)
    stack = []
    unmatched_right = []

    for i, char in enumerate(text_list):
        if char == '（':
            stack.append(i)
        elif char == '）':
            if stack:
                stack.pop()
            else:
                unmatched_right.append(i)

    for i in reversed(unmatched_right):
        del text_list[i]
        issues.append("删除多余右括号1个")

    if stack:
        text_list.extend(['）'] * len(stack))
        issues.append(f"补充缺失右括号{len(stack)}个")

    text = ''.join(text_list)

    # 嵌套修正
    text, nested_count = NESTED_PAREN_PATTERN.subn(r'（\1）', text)
    if nested_count > 0:
        issues.append(f"修复嵌套括号{nested_count}处")

    # ========== 清理空括号或纯标点括号 ==========
    def clean_empty_paren(m):
        content = m.group(1).strip('，、,;；:：。！？.!? ')
        if not content:
            issues.append("删除空括号或仅含标点括号")
            return ''
        return f'（{content}）'

    text = re.sub(r'（(.*?)）', clean_empty_paren, text)

    # ========== 去重 ==========
    seen = set()

    def dedup(m):
        c = m.group(1)
        if c in seen:
            issues.append(f"重复括号内容：'{c}'")
            return ''
        seen.add(c)
        return f'（{c}）'

    text = re.sub(r'（(.*?)）', dedup, text)

    # ========== 多余标点简化 ==========
    text = REGEX_PATTERNS['excess_punct'].sub(lambda m: m.group(0)[0], text)

    # ========== 错别字修正 ==========
    for typo, corr in TYPO_DICT.items():
        if typo in text:
            text = text.replace(typo, corr)
            issues.append(f"错别字：'{typo}'→'{corr}'")

    return text, issues


def process_chunk(chunk):
    """
    处理数据块。支持上传文件列名与导出列名并存：
    学校名称/院校名称、招生专业/专业名称、招生科类/科类、选科要求/报考要求。
    选科转换逻辑与 docx 一致：不限/单字/且/或 → 选科要求说明、次选。
    """
    # 学校名称检查（支持 学校名称 或 院校名称）
    school_col = '学校名称' if '学校名称' in chunk.columns else ('院校名称' if '院校名称' in chunk.columns else None)
    if school_col:
        chunk['学校匹配结果'] = chunk[school_col].apply(check_school_name)

    # 专业匹配检查（支持 招生专业 或 专业名称，需有一级层次）
    major_col = '招生专业' if '招生专业' in chunk.columns else ('专业名称' if '专业名称' in chunk.columns else None)
    if major_col and '一级层次' in chunk.columns:
        chunk['招生专业匹配结果'] = chunk.apply(
            lambda r: check_major_combo(r[major_col], r['一级层次']), axis=1)

    # 备注处理（支持 专业备注）
    remark_col = None
    for c in chunk.columns:
        if '专业备注' in str(c):
            remark_col = c
            break
    if remark_col is not None:
        def process_remark(remark):
            if pd.isna(remark) or not str(remark).strip():
                return '无问题', ''
            fixed_text, issues = analyze_and_fix(remark)
            return '；'.join(issues) if issues else '无问题', fixed_text

        chunk[['备注检查结果', '修改后备注']] = chunk[remark_col].apply(
            lambda x: pd.Series(process_remark(x)))

    # 分数检查
    score_columns = ['最高分', '平均分', '最低分']
    if all(col in chunk.columns for col in score_columns):
        chunk['分数检查结果'] = chunk.apply(check_score_consistency, axis=1)

    # 选科要求处理：依据 docx，支持 选科要求 或 报考要求，统一用 convert_selection_requirement_from_requirement
    req_col = '选科要求' if '选科要求' in chunk.columns else ('报考要求' if '报考要求' in chunk.columns else None)
    if req_col:
        chunk[['选科要求说明', '次选']] = chunk[req_col].apply(
            lambda x: pd.Series(convert_selection_requirement_from_requirement(x)))

    # 招生科类处理（支持 招生科类 或 科类），统一为物理类/历史类并生成首选科目
    cat_col = '招生科类' if '招生科类' in chunk.columns else ('科类' if '科类' in chunk.columns else None)
    if cat_col:
        chunk['招生科类'] = chunk[cat_col].replace({'物理': '物理类', '历史': '历史类'})
        chunk['首选科目'] = chunk['招生科类'].apply(
            lambda x: _get_first_subject(x) if pd.notna(x) and str(x).strip() else '')
    elif '首选科目' not in chunk.columns and req_col:
        chunk['首选科目'] = ''

    return chunk


# ============================
# 院校分提取相关函数（普通类）
# ============================
expected_columns = [
    '学校名称', '省份', '招生专业', '专业方向（选填）', '专业备注（选填）', '一级层次', '招生科类', '招生批次',
    '招生类型（选填）', '最高分', '最低分', '平均分', '最低分位次（选填）', '招生人数（选填）', '数据来源',
    '专业组代码', '首选科目', '选科要求', '次选科目', '专业代码', '招生代码', '录取人数（选填）'
]
columns_to_convert = [
    '专业组代码', '专业代码', '招生代码', '最高分', '最低分', '最低分位次（选填）',
    '招生人数（选填）'
]


def process_score_file(file_path):
    # 首先读取年份（从B2单元格）
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        year_value = ws['B2'].value
        if year_value is None:
            # 如果B2为空，尝试从数据中提取年份
            year_value = ''
        else:
            year_value = str(year_value).strip()
        wb.close()
    except Exception as e:
        year_value = ''

    try:
        df = pd.read_excel(file_path, header=2, dtype={
            '专业组代码': str,
            '专业代码': str,
            '招生代码': str,
            '最高分': str,
            '最低分': str,
            '最低分位次（选填）': str,
            '招生人数（选填）': str,
            '录取人数（选填）': str
        }, keep_default_na=False, engine='openpyxl')
    except Exception as e:
        raise Exception(f"读取文件错误：{e}")

    missing_columns = [col for col in expected_columns if col not in df.columns]
    if missing_columns:
        raise Exception(f"文件缺少以下列：{missing_columns}")

    df['最低分'] = pd.to_numeric(df['最低分'], errors='coerce')
    df['最高分'] = pd.to_numeric(df['最高分'], errors='coerce')
    df['招生人数（选填）'] = pd.to_numeric(df['招生人数（选填）'], errors='coerce')
    df['录取人数（选填）'] = pd.to_numeric(df['录取人数（选填）'], errors='coerce')
    df = df.dropna(subset=['最低分'])

    if df.empty:
        raise Exception("数据处理后为空。")

    df['招生类型（选填）'] = df['招生类型（选填）'].fillna('')

    # 首选科目转换逻辑
    if '首选科目' in df.columns:
        df['首选科目'] = df['首选科目'].str.strip()  # 去除前后空格
        df['首选科目'] = df['首选科目'].replace({
            '历': '历史',
            '物': '物理',
            '历史': '历史',  # 确保已经是"历史"的不变
            '物理': '物理'  # 确保已经是"物理"的不变
        })

    try:
        # 判断是否有专业组代码列，且不全为空
        if '专业组代码' in df.columns and df['专业组代码'].notna().any():
            group_fields = ['学校名称', '省份', '一级层次', '招生科类', '招生批次', '招生类型（选填）', '专业组代码']
        else:
            group_fields = ['学校名称', '省份', '一级层次', '招生科类', '招生批次', '招生类型（选填）']

        # 每组最低分所在行
        min_indices = df.groupby(group_fields)['最低分'].idxmin()

        # 每组最高分
        max_scores = df.groupby(group_fields)['最高分'].max()

        # 取最低分行
        result = df.loc[min_indices].copy()

        # 补充最高分
        def get_max_score(row):
            key = tuple(row[col] for col in group_fields)
            return max_scores.get(key, None)

        result['最高分'] = result.apply(get_max_score, axis=1)

        # 招生人数、录取人数按分组总和
        enroll_groups = df.groupby(group_fields)['招生人数（选填）'].sum()
        code_groups = df.groupby(group_fields)['录取人数（选填）'].sum()

        def get_group_total(row, column_name):
            key = tuple(row[col] for col in group_fields)
            if column_name == '招生人数（选填）':
                return enroll_groups.get(key, '')
            elif column_name == '录取人数（选填）':
                return code_groups.get(key, '')
            return ''

        result['招生人数（选填）'] = result.apply(lambda row: get_group_total(row, '招生人数（选填）'), axis=1)
        result['录取人数（选填）'] = result.apply(lambda row: get_group_total(row, '录取人数（选填）'), axis=1)

    except Exception as e:
        raise Exception(f"分组字段错误：{e}")

    if result.empty:
        raise Exception("筛选结果为空。")

    # 构建新的数据框，按照新的列顺序
    new_columns = [
        '学校名称', '省份', '招生类别', '招生批次', '招生类型', '选测等级',
        '最高分', '最低分', '平均分', '最高位次', '最低位次', '平均位次',
        '录取人数', '招生人数', '数据来源', '省控线科类', '省控线批次', '省控线备注',
        '专业组代码', '首选科目', '院校招生代码'
    ]

    # 创建新的DataFrame，确保所有列都有正确的长度
    num_rows = len(result)
    new_result = pd.DataFrame(index=range(num_rows))

    # 辅助函数：处理列值，将NaN转换为空字符串（用于文本列）
    def get_col_values(col_name, default=''):
        if col_name in result.columns:
            values = result[col_name].fillna(default).astype(str).values
            # 将'nan'字符串转换回空字符串
            values = ['' if str(v).lower() == 'nan' else v for v in values]
            return values
        else:
            return [default] * num_rows

    # 辅助函数：处理数字列值，保持数字类型
    def get_numeric_values(col_name, default=0):
        if col_name in result.columns:
            values = result[col_name].fillna(default)
            # 尝试转换为数字，无法转换的保持原值或设为默认值
            try:
                return pd.to_numeric(values, errors='coerce').fillna(default).values
            except:
                return [default] * num_rows
        else:
            return [default] * num_rows

    new_result['学校名称'] = get_col_values('学校名称')
    new_result['省份'] = get_col_values('省份')
    new_result['招生类别'] = get_col_values('招生科类')
    new_result['招生批次'] = get_col_values('招生批次')
    new_result['招生类型'] = get_col_values('招生类型（选填）')
    new_result['选测等级'] = [''] * num_rows  # 新字段，设为空
    new_result['最高分'] = get_col_values('最高分')
    new_result['最低分'] = get_col_values('最低分')
    new_result['平均分'] = [''] * num_rows  # 删除平均分提取逻辑，设为空
    new_result['最高位次'] = [''] * num_rows  # 新字段，设为空
    new_result['最低位次'] = get_col_values('最低分位次（选填）')
    new_result['平均位次'] = [''] * num_rows  # 新字段，设为空
    new_result['录取人数'] = get_numeric_values('录取人数（选填）', default=0)  # 保持数字格式
    new_result['招生人数'] = get_numeric_values('招生人数（选填）', default=0)  # 保持数字格式
    new_result['数据来源'] = get_col_values('数据来源')
    new_result['省控线科类'] = [''] * num_rows  # 新字段，设为空
    new_result['省控线批次'] = [''] * num_rows  # 新字段，设为空
    new_result['省控线备注'] = [''] * num_rows  # 新字段，设为空
    new_result['专业组代码'] = get_col_values('专业组代码')
    new_result['首选科目'] = get_col_values('首选科目')
    new_result['院校招生代码'] = get_col_values('招生代码')

    output_path = file_path.replace('.xlsx', '_院校分.xlsx')

    try:
        # 创建备注文本
        remark_text = """备注：请删除示例后再填写；
1.省份：必须填写各省份简称，例如：北京、内蒙古，不能带有市、省、自治区、空格、特殊字符等
2.科类：浙江、上海限定"综合、艺术类、体育类"，内蒙古限定"文科、理科、蒙授文科、蒙授理科、艺术类、艺术文、艺术理、体育类、体育文、体育理、蒙授艺术、蒙授体育"，其他省份限定"文科、理科、艺术类、艺术文、艺术理、体育类、体育文、体育理"
3.批次：（以下为19年使用批次）
    北京、天津、辽宁、上海、山东、广东、海南限定本科提前批、本科批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
    河北、内蒙古、吉林、江苏、安徽、福建、江西、河南、湖北、广西、重庆、四川、贵州、云南、西藏、陕西、甘肃、宁夏、新疆限定本科提前批、本科一批、本科二批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
    黑龙江、湖南、青海限定本科提前批、本科一批、本科二批、本科三批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
    山西限定本科一批A段、本科一批B段、本科二批A段、本科二批B段、本科二批C段、专科批、国家专项计划本科批、地方专项计划本科批；
    浙江限定普通类提前批、平行录取一段、平行录取二段、平行录取三段
4.最高分、最低分、平均分：仅能填写数字（最多保留2位小数），且三者顺序不能改变，最低分为必填项，其中艺术类和体育类分数为文化课分数
5.最低分位次：仅能填写数字
6.录取人数：仅能填写数字
7.首选科目：新八省必填，只能填写（历史或物理）"""

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 先写入数据（不包含标题，从第4行开始）
            new_result.to_excel(writer, index=False, header=False, startrow=3)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # 第一行：合并A1-U1并写入备注
            worksheet.merge_cells('A1:U1')
            worksheet['A1'] = remark_text
            worksheet['A1'].alignment = Alignment(wrap_text=True, vertical='top')
            # 设置第一行行高为215磅
            worksheet.row_dimensions[1].height = 215

            # 第二行：A2="招生年"，B2=年份，C2="1"，D2="模板类型（模板标识不要更改）"
            worksheet['A2'] = '招生年'
            # B2和C2设置为数字格式
            try:
                # 尝试将年份转换为数字
                if year_value and str(year_value).strip():
                    year_num = int(float(str(year_value).strip()))
                    worksheet['B2'] = year_num
                else:
                    worksheet['B2'] = ''
            except:
                worksheet['B2'] = year_value
            worksheet['C2'] = 1  # 直接设置为数字1
            worksheet['D2'] = '模板类型（模板标识不要更改）'

            # 第三行：标题行
            headers = ['学校名称', '省份', '招生类别', '招生批次', '招生类型', '选测等级',
                       '最高分', '最低分', '平均分', '最高位次', '最低位次', '平均位次',
                       '录取人数', '招生人数', '数据来源', '省控线科类', '省控线批次', '省控线备注',
                       '专业组代码', '首选科目', '院校招生代码']
            for col_idx, header in enumerate(headers, start=1):
                worksheet.cell(row=3, column=col_idx, value=header)

            # 设置文本格式（从第4行开始，即数据行）
            # 需要设置为文本格式的列（使用新列名，不包括招生人数和录取人数）
            text_format_cols = ['专业组代码', '院校招生代码', '最高分', '最低分', '最低位次']
            for col in text_format_cols:
                if col in new_result.columns:
                    col_idx = new_result.columns.get_loc(col) + 1
                    for row in range(4, len(new_result) + 4):
                        worksheet.cell(row=row, column=col_idx).number_format = numbers.FORMAT_TEXT

            # 确保B2和C2单元格保持数字格式
            if worksheet['B2'].value is not None and str(worksheet['B2'].value).strip():
                try:
                    worksheet['B2'].value = int(float(str(worksheet['B2'].value)))
                except:
                    pass
            worksheet['C2'].value = 1

            # 确保"录取人数"和"招生人数"列保持数字格式（从第4行开始）
            if '录取人数' in new_result.columns:
                col_idx = new_result.columns.get_loc('录取人数') + 1
                for row in range(4, len(new_result) + 4):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) if str(cell.value).strip() else 0
                        except:
                            pass

            if '招生人数' in new_result.columns:
                col_idx = new_result.columns.get_loc('招生人数') + 1
                for row in range(4, len(new_result) + 4):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) if str(cell.value).strip() else 0
                        except:
                            pass

        return output_path
    except Exception as e:
        raise Exception(f"文件保存失败：{e}")


# ============================
# 保持文本格式
# ============================
def _find_remark_column(df):
    """在 DataFrame 中查找专业备注相关列（上传多为“专业备注”，新文件多为“专业备注（选填）”）"""
    for col in df.columns:
        c = str(col).strip() if col is not None else ""
        if not c:
            continue
        if c in ("专业备注", "专业备注（选填）") or "专业备注" in c:
            return col
    return None


def process_remarks_file(file_path, progress_callback=None):
    """学业桥数据处理：上传文件第1行为标题，校验指定列；校对学校/专业/备注后按新格式导出。"""
    try:
        # 上传文件从第一行（标题行）开始读取
        df = pd.read_excel(file_path, header=0, dtype={
            '招生代码': str,
            '专业组编号': str,
            '专业代码': str,
        }, engine='openpyxl', keep_default_na=False)
    except Exception as e:
        raise Exception(f"读取文件错误：{e}")
    # 校验必须包含的列（学业桥上传格式）
    missing = [c for c in XUEYEQIAO_UPLOAD_COLUMNS if c not in df.columns]
    if missing:
        raise Exception("上传文件缺少以下列（应从第1行标题开始）：%s。当前列名：%s" % (missing, list(df.columns)))
    for col in ['招生代码', '专业组编号', '专业代码']:
        if col in df.columns:
            df[col] = df[col].astype(str)
    # 专业备注列已在上传列中，无需再查找或重命名
    chunks = []
    for i in range(0, len(df), 1000):
        chunks.append(df.iloc[i:i + 1000].copy())
    results = {}
    total_chunks = len(chunks)
    with ThreadPoolExecutor(max_workers=os.cpu_count() or 4) as executor:
        future_to_index = {executor.submit(process_chunk, chunk): idx for idx, chunk in enumerate(chunks)}
        for count, future in enumerate(as_completed(future_to_index)):
            idx = future_to_index[future]
            results[idx] = future.result()
            if progress_callback:
                progress_callback(count + 1, total_chunks)
    ordered_results = [results[i] for i in sorted(results.keys())]
    final_result = pd.concat(ordered_results)
    # 从上传数据取招生年份（年份列第一个非空值）
    year_value = ''
    if '年份' in final_result.columns:
        for v in final_result['年份']:
            if pd.notna(v) and str(v).strip():
                year_value = str(v).strip()
                break
    # 将每一行映射为导出格式（含 process_chunk 产生的修改后备注等）
    export_rows = []
    for _, row in final_result.iterrows():
        export_rows.append(map_upload_row_to_export(row.to_dict()))
    export_df = pd.DataFrame(export_rows, columns=XUEYEQIAO_EXPORT_HEADERS)
    # 最高分、最低分、平均分：仅数字保留小数后两位
    def _format_score(x):
        if x is None or (isinstance(x, str) and not x.strip()):
            return ''
        s = str(x).strip()
        if not s or not _is_numeric_str(s):
            return s
        return '%.2f' % float(s)
    for col in ['最高分', '最低分', '平均分']:
        if col in export_df.columns:
            export_df[col] = export_df[col].apply(_format_score)
    output_path = file_path.replace('.xlsx', '_检查结果.xlsx')
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        # 第1行：A1-U1 合并，行高 220 磅，备注内容
        ws.merge_cells('A1:U1')
        ws['A1'] = XUEYEQIAO_EXPORT_NOTE
        ws['A1'].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        ws.row_dimensions[1].height = 220
        # 第2行：A2=招生年份，B2=年份
        ws['A2'] = '招生年份'
        ws['B2'] = year_value
        # 第3行：标题行
        for col_idx, col_name in enumerate(XUEYEQIAO_EXPORT_HEADERS, start=1):
            ws.cell(row=3, column=col_idx, value=col_name)
        # 第4行起：数据
        for row_idx, (_, row_data) in enumerate(export_df.iterrows(), start=4):
            for col_idx, col_name in enumerate(XUEYEQIAO_EXPORT_HEADERS, start=1):
                val = row_data.get(col_name)
                if pd.isna(val):
                    val = ''
                ws.cell(row=row_idx, column=col_idx, value=val)
        # 专业组代码、专业代码、招生代码等列为文本格式
        text_cols = ['专业组代码', '专业代码', '招生代码', '最低分位次（选填）', '招生人数（选填）', '最低分数区间低', '最低分数区间高', '最低分数区间位次低', '最低分数区间位次高', '录取人数（选填）']
        for col_name in text_cols:
            if col_name in XUEYEQIAO_EXPORT_HEADERS:
                col_idx = XUEYEQIAO_EXPORT_HEADERS.index(col_name) + 1
                for r in range(4, len(export_df) + 4):
                    cell = ws.cell(row=r, column=col_idx)
                    if cell.value is not None and str(cell.value).strip() != '':
                        cell.number_format = numbers.FORMAT_TEXT
        wb.save(output_path)
    except Exception as e:
        raise Exception(f"保存文件错误：{e}")
    return output_path


def _is_numeric_str(s):
    """判断字符串是否为数字（含小数）"""
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


# ============================
# 院校分数据处理（艺体类）
# ============================

expected_new_columns = [
    '学校名称', '省份', '专业', '专业方向（选填）', '专业备注（选填）', '专业层次',
    '专业类别', '是否校考', '招生类别', '招生批次', '最低分', '最低分位次（选填）',
    '专业组代码', '首选科目', '选科要求', '次选科目', '招生代码', '校统考分',
    '校文化分', '专业代码', '数据来源'
]
columns_to_convert_new = [
    '专业组代码', '专业代码', '招生代码', '最低分', '最低分位次（选填）',
    '校统考分', '校文化分'
]


def process_new_template_file(file_path):
    # 首先读取原始文件的B2单元格内容
    try:
        wb_original = openpyxl.load_workbook(file_path, data_only=True)
        ws_original = wb_original.active
        b2_value = ws_original['B2'].value
        if b2_value is None:
            b2_value = ''
        else:
            b2_value = str(b2_value).strip()
        wb_original.close()
    except Exception as e:
        b2_value = ''

    try:
        df = pd.read_excel(file_path, header=2, dtype={
            '专业组代码': str,
            '专业代码': str,
            '招生代码': str,
            '最低分': str,
            '最低分位次（选填）': str,
            '校统考分': str,
            '校文化分': str
        }, keep_default_na=False, engine='openpyxl')
    except Exception as e:
        raise Exception(f"读取文件错误：{e}")

    # 检查必需列
    missing_columns = [col for col in expected_new_columns if col not in df.columns]
    if missing_columns:
        raise Exception(f"文件缺少以下列：{missing_columns}")

    # 数值列转为数值型
    df['最低分'] = pd.to_numeric(df['最低分'], errors='coerce')
    df['校统考分'] = pd.to_numeric(df['校统考分'], errors='coerce')
    df['校文化分'] = pd.to_numeric(df['校文化分'], errors='coerce')

    # 删除最低分为空的行
    df = df.dropna(subset=['最低分'])
    if df.empty:
        raise Exception("数据处理后为空。")

    # 首选科目清洗
    if '首选科目' in df.columns:
        df['首选科目'] = df['首选科目'].str.strip()
        df['首选科目'] = df['首选科目'].replace({
            '历': '历史',
            '物': '物理',
            '历史': '历史',
            '物理': '物理'
        })

    try:
        # 判断分组字段
        if '专业组代码' in df.columns and df['专业组代码'].notna().any():
            group_fields = ['学校名称', '省份', '专业方向（选填）', '专业层次', '专业类别', '招生类别', '招生批次',
                            '专业组代码']
        else:
            group_fields = ['学校名称', '省份', '专业方向（选填）', '专业层次', '专业类别', '招生类别', '招生批次']

        # 每组最低分所在行
        min_indices = df.groupby(group_fields)['最低分'].idxmin()

        # 取最低分行
        result = df.loc[min_indices].copy()

    except Exception as e:
        raise Exception(f"分组字段错误：{e}")

    if result.empty:
        raise Exception("筛选结果为空。")

    # 准备新的列名映射
    new_columns = ['学校名称', '省份', '招生类别', '招生批次', '专业类别', '投档分', '位次', '招生代码', '专业组', '备注', '是否校考']
    
    # 创建新的DataFrame，映射字段
    new_result = pd.DataFrame()
    new_result['学校名称'] = result['学校名称'] if '学校名称' in result.columns else pd.Series([None] * len(result))
    new_result['省份'] = result['省份'] if '省份' in result.columns else pd.Series([None] * len(result))
    new_result['招生类别'] = result['招生类别'] if '招生类别' in result.columns else pd.Series([None] * len(result))
    new_result['招生批次'] = result['招生批次'] if '招生批次' in result.columns else pd.Series([None] * len(result))
    new_result['专业类别'] = result['专业类别'] if '专业类别' in result.columns else pd.Series([None] * len(result))
    new_result['投档分'] = result['最低分'] if '最低分' in result.columns else pd.Series([None] * len(result))
    new_result['位次'] = result['最低分位次（选填）'] if '最低分位次（选填）' in result.columns else pd.Series([None] * len(result))
    new_result['招生代码'] = result['招生代码'] if '招生代码' in result.columns else pd.Series([None] * len(result))
    new_result['专业组'] = result['专业组代码'] if '专业组代码' in result.columns else pd.Series([None] * len(result))
    new_result['备注'] = result['专业备注（选填）'] if '专业备注（选填）' in result.columns else pd.Series([None] * len(result))
    # 是否校考：如果存在则使用，否则默认为'否'
    if '是否校考' in result.columns:
        new_result['是否校考'] = result['是否校考'].fillna('否')
    else:
        new_result['是否校考'] = '否'

    # 输出文件路径
    output_path = file_path.replace('.xlsx', '_院校分.xlsx')

    try:
        # 创建新的工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        # 第一行：A1-K1合并单元格，行高90磅
        ws.merge_cells('A1:K1')
        cell_a1 = ws['A1']
        cell_a1.value = '备注：请删除示例后再填写；\n1.省份：必须填写各省份简称，例如：北京、内蒙古，不能带有市、省、自治区、空格、特殊字符等\n2.最低分位次：仅能填写数字\n3.录取人数：仅能填写数字\n4.是否校考：有效值【是，否】，不填写或不在有效值中默认\'否\''
        cell_a1.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        ws.row_dimensions[1].height = 90

        # 第二行：A2="招生年"，B2=原始文件B2的内容
        ws['A2'] = '招生年'
        ws['B2'] = b2_value

        # 第三行：标题行
        for col_idx, col_name in enumerate(new_columns, start=1):
            ws.cell(row=3, column=col_idx, value=col_name)

        # 第四行开始：数据行
        for row_idx, (_, row_data) in enumerate(new_result.iterrows(), start=4):
            ws.cell(row=row_idx, column=1, value=row_data['学校名称'] if pd.notna(row_data['学校名称']) else None)
            ws.cell(row=row_idx, column=2, value=row_data['省份'] if pd.notna(row_data['省份']) else None)
            ws.cell(row=row_idx, column=3, value=row_data['招生类别'] if pd.notna(row_data['招生类别']) else None)
            ws.cell(row=row_idx, column=4, value=row_data['招生批次'] if pd.notna(row_data['招生批次']) else None)
            ws.cell(row=row_idx, column=5, value=row_data['专业类别'] if pd.notna(row_data['专业类别']) else None)
            ws.cell(row=row_idx, column=6, value=row_data['投档分'] if pd.notna(row_data['投档分']) else None)
            ws.cell(row=row_idx, column=7, value=row_data['位次'] if pd.notna(row_data['位次']) else None)
            ws.cell(row=row_idx, column=8, value=row_data['招生代码'] if pd.notna(row_data['招生代码']) else None)
            ws.cell(row=row_idx, column=9, value=row_data['专业组'] if pd.notna(row_data['专业组']) else None)
            ws.cell(row=row_idx, column=10, value=row_data['备注'] if pd.notna(row_data['备注']) else None)
            ws.cell(row=row_idx, column=11, value=row_data['是否校考'] if pd.notna(row_data['是否校考']) else '否')

        # 设置文本格式（从第4行开始，即数据行）
        # 需要设置为文本格式的列
        text_format_cols = ['招生代码', '专业组', '位次']
        for col_name in text_format_cols:
            col_idx = new_columns.index(col_name) + 1
            for row in range(4, len(new_result) + 4):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    # 将值转换为字符串，然后设置为文本格式
                    cell.value = str(cell.value)
                    cell.number_format = numbers.FORMAT_TEXT

        # 保存文件
        wb.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"文件保存失败：{e}")


# ============================
# 一分一段数据处理
# ============================

def process_segmentation_file(file_path):
    output_path = os.path.splitext(file_path)[0] + "_校验结果.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    ws['E7'] = '累计人数校验结果'
    ws['F7'] = '分数校验结果'
    ws['F2'] = '年份校验'

    # 校验 B2 是否为 2025
    if ws['B2'].value != 2025:
        ws['G2'] = f"× 应为2025，当前为：{ws['B2'].value}"
    else:
        ws['G2'] = "√"

    region = ws['B3'].value
    suffix = "-750"
    if region == "上海":
        suffix = "-660"
    elif region == "海南":
        suffix = "-900"

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # ---------- 第8行特殊处理 ----------
    row = 8
    curr_score = ws[f"A{row}"].value
    curr_num = ws[f"B{row}"].value
    curr_total = ws[f"C{row}"].value

    try:
        score_int = int(float(str(curr_score).split('-')[0]))
    except:
        score_int = None

    inserted = False
    if curr_total is not None:
        if curr_num is None or curr_num == "":
            # 没有人数 → 自动计算
            if row == 8:
                ws[f"B{row}"] = curr_total
            else:
                prev_total = ws[f"C{row - 1}"].value
                if prev_total is not None:
                    ws[f"B{row}"] = curr_total - prev_total
        else:
            # 有人数和累计人数不一致时插入补断点行
            if curr_num != curr_total:
                try:
                    insert_score = score_int + 1
                    insert_num = curr_total - curr_num
                    ws.insert_rows(row)
                    ws[f"A{row}"] = f"{insert_score}{suffix}"  # ✅ 仅加后缀在新增行
                    ws[f"B{row}"] = insert_num
                    ws[f"C{row}"] = insert_num
                    for col in ['A', 'B', 'C', 'E', 'F']:
                        ws[f"{col}{row}"].fill = yellow_fill
                    ws[f"E{row}"] = "补断点"
                    ws[f"F{row}"] = "补断点"
                    inserted = True
                except:
                    pass

    # 仅当没有插入行时，第8行加后缀
    if not inserted and score_int is not None:
        ws[f"A{row}"] = f"{score_int}{suffix}"

    # ---------- 补断点逻辑 ----------
    while row < ws.max_row:
        curr = ws[f"A{row}"].value
        next = ws[f"A{row + 1}"].value
        try:
            curr_score_int = int(str(curr).split('-')[0])
            next_score_int = int(str(next).split('-')[0])
        except:
            row += 1
            continue

        if curr_score_int - next_score_int > 1:
            missing_score = curr_score_int - 1
            ws.insert_rows(row + 1)
            ws[f"A{row + 1}"] = missing_score
            ws[f"B{row + 1}"] = 0
            ws[f"C{row + 1}"] = ws[f"C{row}"].value
            for col in ['A', 'B', 'C', 'E', 'F']:
                ws[f"{col}{row + 1}"].fill = yellow_fill
            ws[f"E{row + 1}"] = "补断点"
            ws[f"F{row + 1}"] = "补断点"
        else:
            row += 1

    # ---------- 校验与自动补人数 ----------
    for row in range(8, ws.max_row + 1):
        curr_score = ws[f"A{row}"].value
        curr_num = ws[f"B{row}"].value
        curr_total = ws[f"C{row}"].value
        prev_total = ws[f"C{row - 1}"].value if row > 8 else None
        prev_score = ws[f"A{row - 1}"].value if row > 8 else None

        # 自动补人数
        if (curr_num is None or curr_num == "") and curr_total is not None:
            if row == 8:
                ws[f"B{row}"] = curr_total
                curr_num = curr_total
            elif prev_total is not None:
                try:
                    calc = curr_total - prev_total
                    ws[f"B{row}"] = calc
                    curr_num = calc
                except:
                    pass

        # 校验累计人数
        if row == 8:
            # 第8行直接标记正确（假设第8行累计人数正确）
            if ws[f"E{row}"].value != "补断点":
                ws[f"E{row}"] = "√"
            correct_total = curr_total
        else:
            if curr_num is not None and curr_total is not None and correct_total is not None:
                expected_total = correct_total + curr_num
                if expected_total == curr_total:
                    if ws[f"E{row}"].value != "补断点":
                        ws[f"E{row}"] = "√"
                    correct_total = curr_total  # 本行累计正确，用它更新基准
                else:
                    if ws[f"E{row}"].value != "补断点":
                        ws[f"E{row}"] = f"× 应为{expected_total}"
                    correct_total = expected_total

        # 校验分数差
        try:
            curr_score_num = float(str(curr_score).split('-')[0])
            prev_score_num = float(str(prev_score).split('-')[0])
        except:
            curr_score_num = prev_score_num = None

        if curr_score_num is not None and prev_score_num is not None:
            diff = prev_score_num - curr_score_num
            if diff == 1:
                if ws[f"F{row}"].value != "补断点":
                    ws[f"F{row}"] = "√"
            else:
                if ws[f"F{row}"].value != "补断点":
                    ws[f"F{row}"] = f"× 差值{diff}"
        else:
            if ws[f"F{row}"].value != "补断点":
                ws[f"F{row}"] = "× 分数非数字，无法校验"

    wb.save(output_path)
    return output_path


# ============================
# 专业组代码匹配导出函数
# ============================
def export_match_result_to_excel(export_df, headers, year_value, output_path):
    """导出专业组代码匹配结果为Excel格式"""
    # 创建备注文本
    remark_text = """备注：请删除示例后再填写；
1.省份：必须填写各省份简称，例如：北京、内蒙古，不能带有市、省、自治区、空格、特殊字符等2.科类：浙江、上海限定"综合、艺术类、体育类"，内蒙古限定"文科、理科、蒙授文科、蒙授理科、艺术类、艺术文、艺术理、体育类、体育文、
体育理、蒙授艺术、蒙授体育"，其他省份限定"文科、理科、艺术类、艺术文、艺术理、体育类、体育文、体育理"
3.批次：（以下为19年使用批次）
河北、内蒙古、吉林、江苏、安徽、福建、江西、河南、湖北、广西、重庆、四川、贵州、云南、西藏、陕西、甘肃、宁夏、新疆限定本科提前批、
本科一批、本科二批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
黑龙江、湖南、青海限定本科提前批、本科一批、本科二批、本科三批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
山西限定本科一批A段、本科一批B段、本科二批A段、本科二批B段、本科二批C段、专科批、国家专项计划本科批、地方专项计划本科批；
浙江限定普通类提前批、平行录取一段、平行录取二段、平行录取三段
4.招生人数：仅能填写数字
5.最高分、最低分、平均分：仅能填写数字，保留小数后两位，且三者顺序不能改变，最低分为必填项，其中艺术类和体育类分数为文化课分数
6.一级层次：限定"本科、专科（高职）"，该部分为招生专业对应的专业层次
7.最低分位次：仅能填写数字;
8.数据来源：必须限定——官方考试院、大红本数据、学校官网、销售、抓取、圣达信、优志愿、学业桥
9.选科要求：不限科目专业组;多门选考;单科、多科均需选考
10.选科科目必须是科目的简写（物、化、生、历、地、政、技）
                    
11.2020北京、海南，17-19上海仅限制本科专业组代码必填
12.新八省首选科目必须选择（物理或历史）
13.分数区间仅限北京"""

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 第一行：合并A1-U1并写入备注
    ws.merge_cells('A1:U1')
    ws['A1'] = remark_text
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    # 设置第一行行高为220磅
    ws.row_dimensions[1].height = 220

    # 第二行：A2="招生年份"，B2=年份值
    ws['A2'] = '招生年份'
    ws['B2'] = year_value if year_value else ''
    # B2设置为文本格式
    ws['B2'].number_format = numbers.FORMAT_TEXT

    # 处理标题行：如果headers为空或None，使用export_df的列名
    if not headers or len(headers) == 0:
        headers = list(export_df.columns)
    
    # 清理headers中的None值，并去除空字符串
    headers = [h if h is not None else '' for h in headers]
    
    # 按照headers的顺序导出，确保与原始文件A的第3行标题顺序一致
    # 如果headers中的列在export_df中存在，使用export_df的值；否则为空
    final_headers = []
    for h in headers:
        if h and h.strip():  # 非空标题
            final_headers.append(h.strip())
    
    # 添加export_df中存在但headers中没有的列（追加到末尾）
    for col in export_df.columns:
        if col not in final_headers:
            final_headers.append(col)

    # 第三行：标题行（使用处理后的标题）
    for col_idx, header in enumerate(final_headers, start=1):
        ws.cell(row=3, column=col_idx, value=header if header else '')

    # 数据行（从第4行开始）
    for row_idx, (_, row_data) in enumerate(export_df.iterrows(), start=4):
        for col_idx, header in enumerate(final_headers, start=1):
            if header in export_df.columns:
                value = row_data[header]
                # 处理空值
                if value is None or pd.isna(value):
                    value = ''
                elif isinstance(value, str) and value.lower() in ['nan', 'none']:
                    value = ''
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # 设置代码列为文本格式
                if header in ['专业组代码', '专业代码', '招生代码']:
                    cell.number_format = numbers.FORMAT_TEXT
            else:
                ws.cell(row=row_idx, column=col_idx, value='')

    wb.save(output_path)


# ============================
# 专业组代码匹配
# ============================

tableA_fields = [
    "学校名称", "省份", "招生专业", "专业备注（选填）",
    "一级层次", "招生科类", "招生批次", "招生类型（选填）"
]

rename_mapping_B = {
    "学校": "学校名称",
    "省份": "省份",
    "层次": "一级层次",
    "科类": "招生科类",
    "批次": "招生批次",
    "招生类型": "招生类型（选填）",
    "专业": "招生专业",
    "备注": "专业备注（选填）"
}


def process_data(dfA, dfB):
    dfB.rename(columns=rename_mapping_B, inplace=True)

    # 构建组合键（不含备注和招生类型）：学校-省份-层次-科类-批次-专业
    key_fields = [f for f in tableA_fields if f not in ["专业备注（选填）", "招生类型（选填）"]]
    dfA["组合键"] = dfA[key_fields].fillna("").astype(str).apply(
        lambda x: "|".join([str(i).strip() for i in x]), axis=1)
    dfB["组合键"] = dfB[key_fields].fillna("").astype(str).apply(
        lambda x: "|".join([str(i).strip() for i in x]), axis=1)

    # 检查A表和B表中组合键的重复性
    # 统计A表中每个组合键出现的次数
    a_key_counts = dfA["组合键"].value_counts()
    # 统计B表中每个组合键出现的次数
    b_key_counts = dfB["组合键"].value_counts()

    # 找出A表中有重复的组合键（出现次数>1）
    a_duplicate_keys = set(a_key_counts[a_key_counts > 1].index)
    # 找出B表中有重复的组合键（出现次数>1）
    b_duplicate_keys = set(b_key_counts[b_key_counts > 1].index)

    # 构建B表字典：组合键 → 记录列表
    b_dict = dfB.groupby("组合键").apply(lambda x: x.to_dict("records")).to_dict()

    # 存储需要手动补充的记录信息
    manual_fill_records = []

    def get_code(row):
        key = row["组合键"]
        candidates = b_dict.get(key, [])

        # 检查该组合键在A表或B表中是否有重复
        has_duplicate_in_a = key in a_duplicate_keys
        has_duplicate_in_b = key in b_duplicate_keys

        # 如果A表或B表中任何一个有重复，需要手动补充
        if has_duplicate_in_a or has_duplicate_in_b:
            # 返回完整的候选记录列表（包含所有字段信息）
            return None, candidates if candidates else []

        # A表和B表中都没有重复，且B表中只有唯一候选记录，可以直接匹配
        if len(candidates) == 1:
            return candidates[0]["专业组代码"], None

        # 其他情况（无候选记录或多个候选记录）都需要手动补充
        # 返回None和候选记录列表（可能为空）
        return None, candidates if candidates else []

    # 应用匹配逻辑
    results = dfA.apply(get_code, axis=1)
    dfA["专业组代码"] = results.apply(lambda x: x[0] if x[0] is not None else "")
    
    # 收集需要手动补充的记录（包含完整的候选记录信息）
    # 只要专业组代码没匹配到的，都需要手动选择
    for idx, row in dfA.iterrows():
        result = results.iloc[idx]
        matched_code = result[0]  # 匹配到的专业组代码
        candidates = result[1] if result[1] is not None else []
        
        # 如果专业组代码为空（没有匹配到），需要手动补充
        if not matched_code or matched_code == "":
            # 提取候选记录的详细信息
            candidate_records = []
            for candidate in candidates:
                candidate_records.append({
                    "专业组代码": candidate.get("专业组代码", ""),
                    "学校名称": candidate.get("学校名称", ""),
                    "省份": candidate.get("省份", ""),
                    "招生专业": candidate.get("招生专业", ""),
                    "一级层次": candidate.get("一级层次", ""),
                    "招生科类": candidate.get("招生科类", ""),
                    "招生批次": candidate.get("招生批次", ""),
                    "招生类型（选填）": candidate.get("招生类型（选填）", ""),
                    "备注（招生计划）": candidate.get("专业备注（选填）", ""),  # B表重命名后的备注字段
                })
            
            manual_fill_records.append({
                "索引": idx,
                "学校名称": row.get("学校名称", ""),
                "省份": row.get("省份", ""),
                "招生专业": row.get("招生专业", ""),
                "一级层次": row.get("一级层次", ""),
                "招生科类": row.get("招生科类", ""),
                "招生批次": row.get("招生批次", ""),
                "招生类型（选填）": row.get("招生类型（选填）", ""),
                "专业备注（选填）": row.get("专业备注（选填）", ""),  # A表的专业备注字段
                "候选记录": candidate_records  # 完整的候选记录列表（可能为空）
            })

    return dfA, manual_fill_records


# ========== 就业质量报告图片提取 ==========

def fetch_images_static(url, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    image_paths = []
    try:
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        imgs = soup.find_all("img")
        for idx, img in enumerate(imgs, 1):
            src = img.get("src")
            if not src:
                continue
            full_url = urljoin(url, src)
            # 跳过 base64 或 blob 类型
            if full_url.startswith("data:") or full_url.startswith("blob:"):
                continue
            ext = os.path.splitext(urlparse(full_url).path)[1] or ".jpg"
            filename = f"img_{idx:03d}{ext}"
            path = os.path.join(output_folder, filename)
            try:
                img_resp = requests.get(full_url, timeout=10)
                if img_resp.status_code != 200:
                    continue
                content_type = img_resp.headers.get("content-type", "")
                # 仅保存真正的图片类型
                if not content_type.startswith("image/"):
                    continue
                img_data = img_resp.content
                # 验证图片是否可识别
                try:
                    Image.open(io.BytesIO(img_data))
                except Exception:
                    continue
                with open(path, "wb") as f:
                    f.write(img_data)
                image_paths.append(path)
            except Exception:
                continue
    except Exception as e:
        raise Exception(f"静态模式加载失败: {e}")
    return image_paths


def images_to_pdf(image_paths, pdf_path):
    images = []
    for path in sorted(image_paths):
        try:
            img = Image.open(path).convert("RGB")
            images.append(img)
        except Exception:
            continue
    if images:
        images[0].save(pdf_path, save_all=True, append_images=images[1:])
        return True
    return False


# ============================
# 招生计划数据比对与转换工具相关函数
# ============================

def generate_plan_score_key(item):
    """生成招生计划 vs 专业分的组合键"""
    year = str(item.get('年份', '') or '').strip()
    province = str(item.get('省份', '') or '').strip()
    school = str(item.get('学校', '') or '').strip()
    subject = str(item.get('科类', '') or '').strip()
    batch = str(item.get('批次', '') or '').strip()
    major = str(item.get('专业', '') or '').strip()
    level = str(item.get('层次', '') or '').strip()
    group_code = str(item.get('专业组代码', '') or '').strip()
    return f"{year}|{province}|{school}|{subject}|{batch}|{major}|{level}|{group_code}"


def generate_plan_college_key(item):
    """生成招生计划 vs 院校分的组合键，使用新的组合键字段"""
    province = str(item.get('省份', '') or '').strip()
    school = str(item.get('学校', '') or '').strip()
    subject = str(item.get('科类', '') or '').strip()
    batch = str(item.get('批次', '') or '').strip()
    group_code = str(item.get('专业组代码', '') or '').strip()
    recruit_code = str(item.get('招生代码', '') or '').strip()
    return f"{province}|{school}|{subject}|{batch}|{group_code}|{recruit_code}"


def compare_plan_vs_score(plan_df, score_df):
    """比对招生计划 vs 专业分"""
    plan_score_results = []
    score_key_set = set()

    # 为专业分数据建立索引
    for _, item in score_df.iterrows():
        key = generate_plan_score_key(item.to_dict())
        score_key_set.add(key)

    # 比对招生计划数据
    for idx, row in plan_df.iterrows():
        item = row.to_dict()
        key = generate_plan_score_key(item)
        exists = key in score_key_set

        plan_score_results.append({
            'index': idx + 1,
            'originalIndex': idx,
            'keyFields': {
                '年份': item.get('年份', '') or '',
                '省份': item.get('省份', '') or '',
                '学校': item.get('学校', '') or '',
                '科类': item.get('科类', '') or '',
                '批次': item.get('批次', '') or '',
                '专业': item.get('专业', '') or '',
                '层次': item.get('层次', '') or '',
                '专业组代码': item.get('专业组代码', '') or ''
            },
            'exists': exists,
            'otherInfo': {
                '招生人数': item.get('招生人数', '') or '',
                '学费': item.get('学费', '') or '',
                '学制': item.get('学制', '') or '',
                '专业代码': item.get('专业代码', '') or '',
                '招生代码': item.get('招生代码', '') or '',
                '数据来源': item.get('数据来源', '') or '',
                '备注': item.get('备注', '') or '',
                '招生类型': item.get('招生类型', '') or '',
                '专业组选科要求': item.get('专业组选科要求', '') or '',
                '专业选科要求': item.get('专业选科要求(新高考专业省份)', '') or ''
            }
        })

    return plan_score_results


def compare_plan_vs_college(plan_df, college_df):
    """比对招生计划 vs 院校分"""
    plan_college_results = []
    college_key_set = set()

    # 为院校分数据建立索引
    for _, item in college_df.iterrows():
        key = generate_plan_college_key(item.to_dict())
        college_key_set.add(key)

    # 比对招生计划数据
    for idx, row in plan_df.iterrows():
        item = row.to_dict()
        key = generate_plan_college_key(item)
        exists = key in college_key_set

        plan_college_results.append({
            'index': idx + 1,
            'originalIndex': idx,
            'keyFields': {
                '省份': item.get('省份', '') or '',
                '学校': item.get('学校', '') or '',
                '科类': item.get('科类', '') or '',
                '批次': item.get('批次', '') or '',
                '专业组代码': item.get('专业组代码', '') or '',
                '招生代码': item.get('招生代码', '') or ''
            },
            'exists': exists,
            'otherInfo': {
                '年份': item.get('年份', '') or '',
                '专业': item.get('专业', '') or '',
                '层次': item.get('层次', '') or '',
                '招生人数': item.get('招生人数', '') or '',
                '学费': item.get('学费', '') or '',
                '学制': item.get('学制', '') or '',
                '专业代码': item.get('专业代码', '') or '',
                '数据来源': item.get('数据来源', '') or '',
                '备注': item.get('备注', '') or '',
                '招生类型': item.get('招生类型', '') or '',
                '专业组选科要求': item.get('专业组选科要求', '') or '',
                '专业选科要求': item.get('专业选科要求(新高考专业省份)', '') or ''
            }
        })

    return plan_college_results


def filter_unmatched_plan_data_for_college_export(plan_df, college_df):
    """
    过滤出招生计划中不存在于院校分中的数据。
    
    比对逻辑：
    - 按省份、学校、科类、批次、专业组代码、招生代码这几个字段进行比对
    - 只导出招生计划中，这几个字段的组合键不存在的内容
    - 注意：招生计划中可能存在多个相同的组合键，只要院校分存在一个，就不导出
    
    返回：未匹配的招生计划记录列表
    """
    unmatched_records = []
    
    # 为院校分数据建立组合键集合
    college_key_set = set()
    for _, item in college_df.iterrows():
        key = generate_plan_college_key(item.to_dict())
        college_key_set.add(key)
    
    # 遍历招生计划，找出未匹配的记录
    seen_keys = set()  # 记录已处理过的组合键，避免重复
    for idx, row in plan_df.iterrows():
        item = row.to_dict()
        key = generate_plan_college_key(item)
        
        # 只有当组合键不在院校分中，且这个组合键还没有被处理过时，才添加
        if key not in college_key_set and key not in seen_keys:
            seen_keys.add(key)
            unmatched_records.append({
                'index': idx + 1,
                'originalIndex': idx,
                'data': item
            })
    
    return unmatched_records


def get_first_subject(category):
    """获取首选科目：根据招生科类的第一个字"""
    if not category:
        return ''
    category_str = str(category)
    if '物理类' in category_str or '物理' in category_str:
        return '物'
    elif '历史类' in category_str or '历史' in category_str:
        return '历'
    return ''


def convert_level(level):
    """转换层次字段"""
    if not level:
        return ''
    level_str = str(level).lower()
    if '专科' in level_str or '高职' in level_str:
        return '专科(高职)'
    elif '本科' in level_str:
        return '本科(普通)'
    return level


def extract_required_subjects(text):
    """提取必选科目（处理"物化生（3科必选）"格式）"""
    if not text:
        return []

    subjects = []
    subject_map = {
        '物理': '物', '化学': '化', '生物': '生', '历史': '历',
        '地理': '地', '政治': '政', '技术': '技'
    }

    # 清理文本，保留中文和顿号、逗号
    import re
    clean_text = re.sub(r'[^\u4e00-\u9fa5、，,]', '', str(text)).strip()

    # 处理"物化生（3科必选）"格式：直接提取括号前的内容
    if '必选' in text and '（' in text and text.index('必选') > text.index('（'):
        before_bracket = text.split('（')[0]
        clean_text = before_bracket

    # 处理"物、化、生（3科必选）"格式：顿号分隔的科目
    if '、' in clean_text or '，' in clean_text or ',' in clean_text:
        normalized_text = re.sub(r'[、，]', ',', clean_text)
        parts = [p.strip() for p in normalized_text.split(',') if p.strip()]
        for part in parts:
            for full_name, short_name in subject_map.items():
                if full_name in part or part in full_name:
                    if short_name not in subjects:
                        subjects.append(short_name)
                    break
    else:
        # 处理"物化生"这样的连续字符串
        for full_name, short_name in subject_map.items():
            if full_name in clean_text:
                if short_name not in subjects:
                    subjects.append(short_name)

        # 如果没匹配到全名，尝试按字符匹配
        if len(subjects) == 0 and len(clean_text) > 0:
            char_to_short_map = {
                '物': '物', '化': '化', '生': '生', '历': '历',
                '地': '地', '政': '政', '技': '技'
            }
            for char in clean_text:
                if char in char_to_short_map and char_to_short_map[char] not in subjects:
                    subjects.append(char_to_short_map[char])

    return subjects


def extract_required_subjects_with_format(text):
    """提取必选科目（去掉所有标点符号）
    处理格式如：物化生（3科必选）、物、化、生（3科必选）、生、化、物（3科必选）、物化生(3科必选)等
    返回时去掉所有标点符号，只保留科目字符
    """
    if not text:
        return ''
    
    import re
    
    # 处理"物化生（3科必选）"或"物、化、生（3科必选）"或"生、化、物（3科必选）"格式
    # 支持中文括号（、）和英文括号()
    extracted_text = ''
    
    if '必选' in text:
        # 查找所有可能的括号位置
        bracket_patterns = [
            (r'（', r'）'),  # 中文括号
            (r'\(', r'\)'),  # 英文括号
        ]
        
        for left_bracket, right_bracket in bracket_patterns:
            # 查找左括号位置
            left_match = re.search(left_bracket, text)
            if left_match:
                left_pos = left_match.start()
                # 提取括号前的内容
                before_bracket = text[:left_pos].strip()
                if before_bracket:
                    extracted_text = before_bracket
                    break
        
        # 如果没有找到括号，但包含"3科必选"等字样，尝试提取前面的内容
        # 例如："物化生3科必选"或"物、化、生3科必选"
        if not extracted_text and ('3科必选' in text or '三科必选' in text):
            # 找到"必选"的位置
            bi_xuan_pos = text.find('必选')
            if bi_xuan_pos > 0:
                before_bi_xuan = text[:bi_xuan_pos].strip()
                # 移除可能的数字和"科"字
                before_bi_xuan = re.sub(r'\d+科', '', before_bi_xuan).strip()
                if before_bi_xuan:
                    extracted_text = before_bi_xuan
        
        # 去掉所有标点符号（顿号、逗号、空格等），只保留科目字符
        if extracted_text:
            # 只保留科目字符：物、化、生、历、地、政、技等
            subject_chars = ['物', '化', '生', '历', '地', '政', '技']
            cleaned_text = ''.join([char for char in extracted_text if char in subject_chars])
            return cleaned_text
    
    return ''


def convert_selection_requirement(group_requirement, major_requirement):
    """转换选科要求"""
    selection_requirement = ''
    second_subject = ''

    # 合并两个要求字段（专业组选科要求和专业选科要求）
    group_req_str = str(group_requirement).strip() if group_requirement else ''
    major_req_str = str(major_requirement).strip() if major_requirement else ''
    
    # 如果两个字段都有内容，用顿号连接
    if group_req_str and major_req_str:
        requirement = group_req_str + '、' + major_req_str
    else:
        requirement = group_req_str + major_req_str

    # 清理特殊字符
    import re
    requirement = re.sub(r'^\^+', '', requirement).replace('^', '、').strip()

    if not requirement or requirement == '' or requirement == '、':
        return selection_requirement, second_subject

    # 根据附件2示例处理各种情况
    if '不限' in requirement or '再选不限' in requirement:
        selection_requirement = '不限科目专业组'
    elif '必选' in requirement:
        # 对于"3科必选"的情况，提取科目并去掉标点符号
        original_format = extract_required_subjects_with_format(requirement)
        required_subjects = []
        
        if original_format:
            selection_requirement = '单科、多科均需选考'
            second_subject = original_format
        else:
            # 其他必选情况，使用原有逻辑
            required_subjects = extract_required_subjects(requirement)
            if len(required_subjects) > 0:
                selection_requirement = '单科、多科均需选考'
                second_subject = ''.join(required_subjects)

        # 特殊处理：如果包含"首选"，可能需要排除首选科目
        if '首选' in requirement:
            preferred_subjects = []
            if '首选物理' in requirement:
                preferred_subjects.append('物')
            if '首选历史' in requirement:
                preferred_subjects.append('历')
            
            # 如果已经提取了格式（已去掉标点符号），需要从中排除首选科目
            if original_format:
                # 从已去掉标点的字符串中移除首选科目字符
                filtered_format = original_format
                for pref_subj in preferred_subjects:
                    filtered_format = filtered_format.replace(pref_subj, '')
                if filtered_format:
                    second_subject = filtered_format
            elif required_subjects:
                filtered_subjects = [s for s in required_subjects if s not in preferred_subjects]
                if len(filtered_subjects) > 0:
                    second_subject = ''.join(filtered_subjects)
    elif '首选' in requirement and '再选' in requirement:
        re_select_part = requirement.split('再选')[1] if '再选' in requirement else ''
        re_select_subjects = extract_required_subjects(re_select_part)
        if len(re_select_subjects) > 0:
            selection_requirement = '单科、多科均需选考'
            second_subject = ''.join(re_select_subjects)
    elif '或' in requirement or '选1' in requirement:
        subjects = extract_required_subjects(requirement)
        filtered_subjects = [s for s in subjects if s not in ['物', '历']]
        if len(filtered_subjects) > 0:
            selection_requirement = '多门选考'
            second_subject = ''.join(filtered_subjects)
    else:
        subjects = extract_required_subjects(requirement)
        filtered_subjects = [s for s in subjects if s not in ['物', '历']]
        second_subject = ''.join(filtered_subjects)
        if len(filtered_subjects) > 0:
            selection_requirement = '单科、多科均需选考'

    return selection_requirement, second_subject


def convert_to_text(value):
    """转换为文本格式"""
    if not value and value != 0:
        return ''
    text = str(value).lstrip('^').strip()
    if text == '':
        return ''
    text = text.lstrip("'")
    return text


def convert_data(source_data):
    """转换数据主函数"""
    converted = []

    for row in source_data:
        new_row = {}

        # 基础字段映射
        new_row['学校名称'] = row.get('学校', '') or ''
        new_row['省份'] = row.get('省份', '') or ''
        new_row['招生专业'] = row.get('专业', '') or ''
        new_row['招生科类'] = row.get('科类', '') or ''
        new_row['招生批次'] = row.get('批次', '') or ''
        new_row['招生类型（选填）'] = row.get('招生类型', '') or ''
        new_row['专业备注（选填）'] = row.get('备注', '') or ''
        new_row['招生人数（选填）'] = row.get('招生人数', '') or ''
        new_row['数据来源'] = row.get('数据来源', '') or ''

        # 处理层次字段
        new_row['一级层次'] = convert_level(row.get('层次', ''))

        # 处理代码字段（保持文本格式）
        new_row['招生代码'] = convert_to_text(row.get('招生代码', ''))
        new_row['专业代码'] = convert_to_text(row.get('专业代码', ''))
        new_row['专业组代码'] = convert_to_text(row.get('专业组代码', ''))

        # 处理首选科目
        new_row['首选科目'] = get_first_subject(row.get('科类', ''))

        # 处理选科要求
        selection_requirement, second_subject = convert_selection_requirement(
            row.get('专业组选科要求', ''),
            row.get('专业选科要求(新高考专业省份)', '')
        )
        new_row['选科要求'] = selection_requirement
        new_row['次选科目'] = second_subject

        # 其他字段（留空）
        new_row['专业方向（选填）'] = ''
        new_row['最高分'] = ''
        new_row['最低分'] = ''
        new_row['平均分'] = ''
        new_row['最低分位次（选填）'] = ''
        new_row['最低分数区间低'] = ''
        new_row['最低分数区间高'] = ''
        new_row['最低分数区间位次低'] = ''
        new_row['最低分数区间位次高'] = ''
        new_row['录取人数（选填）'] = ''

        converted.append(new_row)

    return converted


def convert_to_college_score_format(conversion_data):
    """将招生计划数据转换为院校分格式"""
    if not conversion_data:
        return []

    # 辅助函数：安全地处理空值，将None、NaN等转换为空字符串
    def safe_str(value, default=''):
        """安全地将值转换为字符串，处理None、NaN等情况"""
        if value is None:
            return default
        if pd.isna(value):
            return default
        value_str = str(value).strip()
        # 检查是否为'nan'、'None'等字符串
        if value_str.lower() in ['nan', 'none', '']:
            return default
        return value_str

    # 构建分组键：省份、学校、科类、批次、招生类型、层次、专业组代码
    # 如果专业组代码为空，则不包含在分组键中
    def get_group_key(item):
        province = safe_str(item.get('省份', ''))
        school = safe_str(item.get('学校', ''))
        subject = safe_str(item.get('科类', ''))
        batch = safe_str(item.get('批次', ''))
        recruit_type = safe_str(item.get('招生类型', ''))
        level = safe_str(item.get('层次', ''))
        group_code = safe_str(item.get('专业组代码', ''))

        # 如果专业组代码为空或只有^，则不包含在分组键中
        if not group_code or group_code == '^' or group_code == '':
            return (province, school, subject, batch, recruit_type, level)
        else:
            return (province, school, subject, batch, recruit_type, level, group_code)

    # 按分组键分组
    grouped_data = {}
    for item in conversion_data:
        key = get_group_key(item)
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(item)

    # 转换为院校分格式
    college_score_data = []
    for key, items in grouped_data.items():
        # 取第一条记录作为基础数据
        base_item = items[0]

        # 计算招生人数总和
        total_recruit_num = 0
        for item in items:
            recruit_num = item.get('招生人数', '') or ''
            if recruit_num and not pd.isna(recruit_num):
                try:
                    total_recruit_num += float(str(recruit_num))
                except:
                    pass

        # 处理专业组代码：如果为空或只有^，则设为空字符串
        group_code = safe_str(base_item.get('专业组代码', '')).lstrip('^')
        if not group_code or group_code == '^':
            group_code = ''

        # 处理院校招生代码：去除开头的^符号
        recruit_code = safe_str(base_item.get('招生代码', '')).lstrip('^')

        # 处理招生人数：保持为字符串格式（文本格式）
        recruit_num_str = str(int(total_recruit_num)) if total_recruit_num > 0 else ''

        # 构建院校分记录
        college_record = {
            '学校名称': safe_str(base_item.get('学校', '')),
            '省份': safe_str(base_item.get('省份', '')),
            '招生类别': safe_str(base_item.get('科类', '')),
            '招生批次': safe_str(base_item.get('批次', '')),
            '招生类型': safe_str(base_item.get('招生类型', '')),
            '选测等级': '',
            '最高分': '',
            '最低分': '',
            '平均分': '',
            '最高位次': '',
            '最低位次': '',
            '平均位次': '',
            '录取人数': '',
            '招生人数': recruit_num_str,
            '数据来源': safe_str(base_item.get('数据来源', '')),
            '省控线科类': '',
            '省控线批次': '',
            '省控线备注': '',
            '专业组代码': group_code,
            '首选科目': '',
            '院校招生代码': recruit_code
        }

        # 处理首选科目：只有招生类别为物理类/历史类时才填入
        category = college_record['招生类别']
        if '物理类' in category or category == '物理':
            college_record['首选科目'] = '物理'
        elif '历史类' in category or category == '历史':
            college_record['首选科目'] = '历史'

        college_score_data.append(college_record)

    return college_score_data


def export_college_score_data_to_excel(college_score_data, conversion_data, output_path):
    """导出院校分格式的Excel文件"""
    # 创建备注文本
    remark_text = """备注：请删除示例后再填写；
1.省份：必须填写各省份简称，例如：北京、内蒙古，不能带有市、省、自治区、空格、特殊字符等
2.科类：浙江、上海限定"综合、艺术类、体育类"，内蒙古限定"文科、理科、蒙授文科、蒙授理科、艺术类、艺术文、艺术理、体育类、体育文、体育理、蒙授艺术、蒙授体育"，其他省份限定"文科、理科、艺术类、艺术文、艺术理、体育类、体育文、体育理"
3.批次：（以下为19年使用批次）
    北京、天津、辽宁、上海、山东、广东、海南限定本科提前批、本科批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
    河北、内蒙古、吉林、江苏、安徽、福建、江西、河南、湖北、广西、重庆、四川、贵州、云南、西藏、陕西、甘肃、宁夏、新疆限定本科提前批、本科一批、本科二批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
    黑龙江、湖南、青海限定本科提前批、本科一批、本科二批、本科三批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
    山西限定本科一批A段、本科一批B段、本科二批A段、本科二批B段、本科二批C段、专科批、国家专项计划本科批、地方专项计划本科批；
    浙江限定普通类提前批、平行录取一段、平行录取二段、平行录取三段
4.最高分、最低分、平均分：仅能填写数字（最多保留2位小数），且三者顺序不能改变，最低分为必填项，其中艺术类和体育类分数为文化课分数
5.最低分位次：仅能填写数字
6.录取人数：仅能填写数字
7.首选科目：新八省必填，只能填写（历史或物理）"""

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 第一行：合并A1-U1并写入备注
    ws.merge_cells('A1:U1')
    ws['A1'] = remark_text
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    # 设置第一行行高为220磅
    ws.row_dimensions[1].height = 220

    # 第二行：A2="招生年"，B2=年份，C2="1"，D2="模板类型（模板标识不要更改）"
    ws['A2'] = '招生年'
    # 从conversion_data中提取年份
    year_value = ''
    if conversion_data and len(conversion_data) > 0:
        year_value = conversion_data[0].get('年份', '') or ''
        if year_value:
            year_value = str(year_value).strip()

    # B2设置为文本格式
    ws['B2'] = year_value
    ws['B2'].number_format = numbers.FORMAT_TEXT
    ws['C2'] = 1
    ws['D2'] = '模板类型（模板标识不要更改）'

    # 第三行：标题行
    headers = ['学校名称', '省份', '招生类别', '招生批次', '招生类型', '选测等级',
               '最高分', '最低分', '平均分', '最高位次', '最低位次', '平均位次',
               '录取人数', '招生人数', '数据来源', '省控线科类', '省控线批次', '省控线备注',
               '专业组代码', '首选科目', '院校招生代码']
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)

    # 数据行（从第4行开始）
    for row_idx, row_data in enumerate(college_score_data, start=4):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, '')

            # 处理空值：将None、NaN、'nan'字符串等转换为空字符串
            if value is None or pd.isna(value):
                value = ''
            elif isinstance(value, str):
                # 检查是否为'nan'、'None'等字符串
                if value.lower() in ['nan', 'none']:
                    value = ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # 设置文本格式的列：招生人数、专业组代码、院校招生代码
            # 这些列需要保持文本格式，即使内容开头为0也不能抹掉
            if header == '专业组代码' or header == '院校招生代码' or header == '招生人数':
                # 确保值为字符串格式，并设置为文本格式
                if value is not None and value != '':
                    cell.value = str(value)
                else:
                    cell.value = ''  # 确保空值写入为空字符串
                cell.number_format = numbers.FORMAT_TEXT

    wb.save(output_path)


def export_converted_data_to_excel(data, conversion_data, output_path):
    """导出转换后的数据为Excel（保持与HTML中相同的格式）"""
    from datetime import datetime

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 第1行：备注（合并单元格）
    remark_text = """备注：请删除示例后再填写；
1.省份：必须填写各省份简称，例如：北京、内蒙古，不能带有市、省、自治区、空格、特殊字符等
2.科类：浙江、上海限定"综合、艺术类、体育类"，内蒙古限定"文科、理科、蒙授文科、蒙授理科、艺术类、艺术文、艺术理、体育类、体育文、体育理、蒙授艺术、蒙授体育"，其他省份限定"文科、理科、艺术类、艺术文、艺术理、体育类、体育文、体育理"
3.批次：（以下为19年使用批次）
河北、内蒙古、吉林、江苏、安徽、福建、江西、河南、湖北、广西、重庆、四川、贵州、云南、西藏、陕西、甘肃、宁夏、新疆限定本科提前批、本科一批、本科二批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
黑龙江、湖南、青海限定本科提前批、本科一批、本科二批、本科三批、专科提前批、专科批、国家专项计划本科批、地方专项计划本科批；
山西限定本科一批A段、本科一批B段、本科二批A段、本科二批B段、本科二批C段、专科批、国家专项计划本科批、地方专项计划本科批；
浙江限定普通类提前批、平行录取一段、平行录取二段、平行录取三段
4.招生人数：仅能填写数字
5.最高分、最低分、平均分：仅能填写数字，保留小数后两位，且三者顺序不能改变，最低分为必填项，其中艺术类和体育类分数为文化课分数
6.一级层次：限定"本科、专科（高职）"，该部分为招生专业对应的专业层次
7.最低分位次：仅能填写数字;
8.数据来源：必须限定——官方考试院、大红本数据、学校官网、销售、抓取、圣达信、优志愿、学业桥
9.选科要求：不限科目专业组;多门选考;单科、多科均需选考
10.选科科目必须是科目的简写（物、化、生、历、地、政、技）

11.2020北京、海南，17-19上海仅限制本科专业组代码必填
12.新八省首选科目必须选择（物理或历史）
13.分数区间仅限北京"""

    ws.merge_cells('A1:Y1')
    ws['A1'] = remark_text
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[1].height = 220

    # 第2行：招生年份
    admission_year = ''
    if conversion_data and len(conversion_data) > 0 and conversion_data[0].get('年份'):
        admission_year = conversion_data[0]['年份']
    ws['A2'] = '招生年份'
    ws['B2'] = admission_year

    # 第3行：表头
    headers = [
        '学校名称', '省份', '招生专业', '专业方向（选填）', '专业备注（选填）',
        '一级层次', '招生科类', '招生批次', '招生类型（选填）', '最高分',
        '最低分', '平均分', '最低分位次（选填）', '招生人数（选填）',
        '数据来源', '专业组代码', '首选科目', '选科要求', '次选科目',
        '专业代码', '招生代码', '最低分数区间低', '最低分数区间高',
        '最低分数区间位次低', '最低分数区间位次高', '录取人数（选填）'
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=header)

    # 数据行
    for row_idx, row_data in enumerate(data, start=4):
        for col_idx, header in enumerate(headers, start=1):
            value = row_data.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # 设置代码列为文本格式
            if header in ['专业组代码', '专业代码', '招生代码']:
                cell.number_format = numbers.FORMAT_TEXT

    # 设置列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 9.36

    wb.save(output_path)


# ============================
# Streamlit页面布局
# ============================
# 页面标题
st.title("📊 数据处理工具")
st.markdown("---")

# 功能说明
with st.expander("📌 功能说明", expanded=True):
    st.markdown("""
    1. 上传的文件使用库中专业分、院校分、招生计划、一分一段的模板，直接上传即可，无需删减
    2. 备注检查中，检查出来括号有问题的内容还需要自己再过一遍；整个文件的备注需要大概看看有没有错别字
    3. 校验一分一段时，内容不能为文本格式
    4. 使用专业组代码匹配时，两份文件中的“学校-省份-层次-科类-批次-类型”这些字段需要保持一致
    """)

# 更新日志对话框
with st.expander("📢 版本更新（2026.1.27更新）（必看！）", expanded=False):
    st.markdown("""
    ### 2026.1.27更新
    • 修改了专业分匹配逻辑（“学校-省份-层次-科类-批次”），重复字段及未匹配到的内容需要手动补充
    • 修改了招生计划数据对比逻辑（需检查无专业组代码的省份的选科要求）

    ### 历史更新

    #### 2025.4.14更新
    • 招生代码和专业代码保持文本格式  
    • 增加功能说明  
    • 优化工具界面  

    #### 2025.4.16更新
    • 优化了院校分提取处理逻辑  

    #### 2025.5.22更新
    • 更新了院校分提取中录取人数的处理逻辑（建议进行抽查）  
    • 学业桥数据处理中增加了最高分、平均分、最低分的校验，会在最后加一列校验结果  

    #### 2025.5.23更新
    • 学业桥数据处理中增加了学校名称和招生专业的匹配  

    #### 2025.5.27更新
    • 学业桥数据处理中，增加了"招生科类"、"首选科目"、"选科要求"，"次选科目"的处理  
      - 学业桥提供的"3+1+2"省份的招生科类为"物理"、"历史"，可以直接转换为标准的"物理类"、"历史类"  
      - "3+1+2"省份的首选科目可以直接根据招生科类提取  
      - 新增了选科要求、次选科目的处理，可直接转换为标准格式，无需手动处理（处理后的数据在文档最后几列）  

    #### 2025.5.30更新
    新增"一分一段数据处理"  
      - 可直接校验分数、累计人数  
      - 自动补断点  
      - 自动增加"最高分——满分"的区间（上海满分660，海南满分900）  

    ### 2025.6.6更新
    "一分一段数据处理"优化  
      - 自动补充"最高分——满分"的区间（上海满分660，海南满分900）  
      - 只有累计人数没有人数时，可计算人数，无需手动操作  
      - 补断点的分数标注颜色，并在分数和人数校验中标注"补断点"

    ### 2025.6.12更新
    院校分提取逻辑更新  
      - 提取最高分改为取同一个“学校-省份-层次-科类-批次-类型（-专业组代码）”下的最高分

    ### 2025.6.14更新
    专业组代码匹配功能  
      - 需要上传专业分导入模板和库中招生计划导出模板
      - 把库中导出招生计划类型尽量补充完整，否则容易出错
      - 匹配结果需要检查

    ### 2025.7.7更新
    就业质量报告图片抓取功能  
      - 抓取就业质量报告图片
      - 如果抓取到的图片比较多，“下载PDF”的弹框会出现比较慢
      - 注意：只能抓取静态页面的图片，动态页面和有限制的网页无法抓取


    ### 2025.9.26更新
    • 更新了院校分中最高分的提取逻辑  
    • 新增了艺体类院校分提取功能，可以直接上传艺体类专业分模板（可把特殊类型<如：中外合作办学>的备注在专业分中放到专业方向再提取）


    """)

# 创建选项卡
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(
    [
        "院校分提取（普通类）",
        "院校分提取（艺体类）",
        "学业桥数据处理",
        "一分一段校验",
        "专业组代码匹配",
        "就业质量报告图片提取",
        "招生计划数据比对"
    ]
)

# ====================== 院校分提取 ======================
with tab1:
    st.header("院校分提取（普通类）")

    # 文件上传
    uploaded_file = st.file_uploader("选择Excel文件", type=["xlsx"], key="score_file")

    if uploaded_file is not None:
        st.success(f"已选择文件: {uploaded_file.name}")

        # 显示处理进度
        progress_bar = st.progress(0)
        status_text = st.empty()
        status_text.text("准备处理...")

        # 处理按钮
        if st.button("开始数据处理", key="process_score"):
            try:
                # 保存上传的文件到临时位置
                temp_file = "temp_score.xlsx"
                with open(temp_file, "wb") as f:
                    f.write(uploaded_file.getbuffer())

                # 处理文件
                for percent_complete in range(0, 101, 10):
                    progress_bar.progress(percent_complete)
                    status_text.text(f"处理中... {percent_complete}%")

                    # 模拟处理过程，实际使用时替换为您的process_score_file函数
                    if percent_complete == 100:
                        output_path = process_score_file(temp_file)

                # 处理完成
                status_text.text("处理完成！")
                st.balloons()

                # 提供下载链接
                with open(output_path, "rb") as f:
                    bytes_data = f.read()
                b64 = base64.b64encode(bytes_data).decode()
                href = f'<a href="data:application/octet-stream;base64,{b64}" download="院校分提取结果.xlsx">点击下载处理结果</a>'
                st.markdown(href, unsafe_allow_html=True)

                # 清理临时文件
                os.remove(temp_file)
                os.remove(output_path)

            except Exception as e:
                st.error(f"处理过程中发生错误: {str(e)}")

# ====================== 院校分提取（艺体类） ======================
with tab2:
    st.header("院校分提取（艺体类）")

    # 文件上传
    uploaded_file_new = st.file_uploader("选择Excel文件", type=["xlsx"], key="new_score_file")

    if uploaded_file_new is not None:
        st.success(f"已选择文件: {uploaded_file_new.name}")

        # 显示处理进度
        progress_bar = st.progress(0)
        status_text = st.empty()
        status_text.text("准备处理...")

        # 处理按钮
        if st.button("开始数据处理", key="process_new_score"):
            try:
                # 保存上传的文件到临时位置
                temp_file = "temp_new_score.xlsx"
                with open(temp_file, "wb") as f:
                    f.write(uploaded_file_new.getbuffer())

                # 处理文件
                for percent_complete in range(0, 101, 10):
                    progress_bar.progress(percent_complete)
                    status_text.text(f"处理中... {percent_complete}%")

                    # 调用新模板处理函数
                    if percent_complete == 100:
                        output_path = process_new_template_file(temp_file)

                # 处理完成
                status_text.text("处理完成！")
                st.balloons()

                # 提供下载链接
                with open(output_path, "rb") as f:
                    bytes_data = f.read()
                b64 = base64.b64encode(bytes_data).decode()
                href = f'<a href="data:application/octet-stream;base64,{b64}" download="院校分（艺体类）提取结果.xlsx">点击下载处理结果</a>'
                st.markdown(href, unsafe_allow_html=True)

                # 清理临时文件
                os.remove(temp_file)
                os.remove(output_path)

            except Exception as e:
                st.error(f"处理过程中发生错误: {str(e)}")

# ====================== 学业桥数据处理 ======================
with tab3:
    st.header("学业桥数据处理")

    # 文件上传
    uploaded_file = st.file_uploader("选择Excel文件", type=["xlsx"], key="remarks_file")

    if uploaded_file is not None:
        st.success(f"已选择文件: {uploaded_file.name}")

        # 显示处理进度
        progress_bar = st.progress(0)
        status_text = st.empty()
        status_text.text("准备处理...")

        # 处理按钮
        if st.button("开始数据处理", key="process_remarks"):
            try:
                # 保存上传的文件到临时位置
                temp_file = "temp_remarks.xlsx"
                with open(temp_file, "wb") as f:
                    f.write(uploaded_file.getbuffer())


                # 进度回调函数
                def update_progress(current, total):
                    percent = int((current / total) * 100)
                    progress_bar.progress(percent)
                    status_text.text(f"处理中... {percent}%")


                # 处理文件
                output_path = process_remarks_file(temp_file, progress_callback=update_progress)

                # 处理完成
                progress_bar.progress(100)
                status_text.text("处理完成！")
                st.balloons()

                # 提供下载链接
                with open(output_path, "rb") as f:
                    bytes_data = f.read()
                b64 = base64.b64encode(bytes_data).decode()
                href = f'<a href="data:application/octet-stream;base64,{b64}" download="学业桥数据处理结果.xlsx">点击下载处理结果</a>'
                st.markdown(href, unsafe_allow_html=True)

                # 清理临时文件
                os.remove(temp_file)
                os.remove(output_path)

            except Exception as e:
                st.error(f"处理过程中发生错误: {str(e)}")

# ====================== 一分一段校验 ======================
with tab4:
    st.header("一分一段校验")

    # 文件上传
    uploaded_file = st.file_uploader("选择Excel文件", type=["xlsx"], key="segmentation_file")

    if uploaded_file is not None:
        st.success(f"已选择文件: {uploaded_file.name}")

        # 显示处理进度
        progress_bar = st.progress(0)
        status_text = st.empty()
        status_text.text("准备处理...")

        # 处理按钮
        if st.button("开始数据处理", key="process_segmentation"):
            try:
                # 保存上传的文件到临时位置
                temp_file = "一分一段.xlsx"
                with open(temp_file, "wb") as f:
                    f.write(uploaded_file.getbuffer())

                # 处理文件
                for percent_complete in range(0, 101, 10):
                    progress_bar.progress(percent_complete)
                    status_text.text(f"处理中... {percent_complete}%")

                    # 模拟处理过程，实际使用时替换为您的process_segmentation_file函数
                    if percent_complete == 100:
                        output_path = process_segmentation_file(temp_file)

                # 处理完成
                status_text.text("处理完成！")
                st.balloons()

                # 提供下载链接
                with open(output_path, "rb") as f:
                    bytes_data = f.read()

                b64 = base64.b64encode(bytes_data).decode()

                # 从 output_path 提取原文件名（去掉扩展名）
                base_name = os.path.splitext(os.path.basename(output_path))[0]

                # 拼接新文件名
                new_filename = f"{base_name}.xlsx"

                # 构造下载链接
                href = f'<a href="data:application/octet-stream;base64,{b64}" download="{new_filename}">点击下载处理结果</a>'

                st.markdown(href, unsafe_allow_html=True)

                # 清理临时文件
                os.remove(temp_file)
                os.remove(output_path)

            except Exception as e:
                st.error(f"处理过程中发生错误: {str(e)}")

# ====================== 专业组代码匹配 ======================
with tab5:
    st.header("专业组代码匹配")

    # 初始化session state
    if 'match_result_df' not in st.session_state:
        st.session_state.match_result_df = None
    if 'manual_fill_records' not in st.session_state:
        st.session_state.manual_fill_records = []
    if 'manual_selections' not in st.session_state:
        st.session_state.manual_selections = {}
    if 'temp_fileA_path' not in st.session_state:
        st.session_state.temp_fileA_path = None
    if 'temp_fileB_path' not in st.session_state:
        st.session_state.temp_fileB_path = None
    if 'fileA_headers' not in st.session_state:
        st.session_state.fileA_headers = None
    if 'fileB_year' not in st.session_state:
        st.session_state.fileB_year = None

    uploaded_fileA = st.file_uploader("上传专业分导入模板", type=["xls", "xlsx"], key="fileA")
    uploaded_fileB = st.file_uploader("上传招生计划数据导出文件", type=["xls", "xlsx"], key="fileB")

    if uploaded_fileA and uploaded_fileB:
        st.success(f"已选择文件：{uploaded_fileA.name} 和 {uploaded_fileB.name}")

        progress_bar = st.progress(0)
        status_text = st.empty()
        status_text.text("等待开始处理...")

        if st.button("开始数据处理", key="start_match"):
            try:
                # 保存临时文件
                temp_fileA = "tempA.xlsx"
                temp_fileB = "tempB.xlsx"
                with open(temp_fileA, "wb") as f:
                    f.write(uploaded_fileA.getbuffer())
                with open(temp_fileB, "wb") as f:
                    f.write(uploaded_fileB.getbuffer())

                st.session_state.temp_fileA_path = temp_fileA
                st.session_state.temp_fileB_path = temp_fileB

                status_text.text("读取文件...")
                progress_bar.progress(10)

                # 读取文件A的标题行（第3行）
                wbA = openpyxl.load_workbook(temp_fileA, data_only=True)
                wsA = wbA.active
                headers_row = []
                # 读取第3行的所有非空单元格
                max_col = wsA.max_column
                for col_idx in range(1, max_col + 1):
                    cell_value = wsA.cell(row=3, column=col_idx).value
                    headers_row.append(cell_value if cell_value is not None else '')
                wbA.close()
                st.session_state.fileA_headers = headers_row

                # 读取文件B的年份（从A列年份字段读取）
                year_value = ''
                try:
                    # 先尝试从数据中读取年份字段（A列）
                    dfB_temp = pd.read_excel(temp_fileB)
                    if '年份' in dfB_temp.columns:
                        year_values = dfB_temp['年份'].dropna()
                        if len(year_values) > 0:
                            year_value = year_values.iloc[0]
                    # 如果年份字段不存在，尝试从A列第一行数据读取
                    elif len(dfB_temp) > 0:
                        # 尝试读取A列的第一行数据
                        first_col = dfB_temp.iloc[:, 0]
                        if len(first_col) > 0:
                            first_value = first_col.iloc[0]
                            # 如果第一行数据看起来像年份（4位数字）
                            if first_value and str(first_value).strip().isdigit() and len(str(first_value).strip()) == 4:
                                year_value = str(first_value).strip()
                    # 如果还是没找到，尝试从Excel文件的A列读取
                    if not year_value or year_value == '':
                        wbB = openpyxl.load_workbook(temp_fileB, data_only=True)
                        wsB = wbB.active
                        # 从A列查找年份（跳过可能的标题行，从第2行开始查找）
                        for row_idx in range(2, min(wsB.max_row + 1, 100)):  # 最多查找100行
                            cell_value = wsB[f'A{row_idx}'].value
                            if cell_value:
                                cell_str = str(cell_value).strip()
                                # 如果看起来像年份（4位数字）
                                if cell_str.isdigit() and len(cell_str) == 4:
                                    year_value = cell_str
                                    break
                        wbB.close()
                    if year_value is not None:
                        year_value = str(year_value).strip()
                    else:
                        year_value = ''
                except Exception as e:
                    logging.warning(f"读取文件B年份失败：{e}")
                    year_value = ''
                st.session_state.fileB_year = year_value

                dfA = pd.read_excel(temp_fileA, header=2)
                dfB = pd.read_excel(temp_fileB)

                status_text.text("开始处理数据...")
                progress_bar.progress(30)

                result_df, manual_fill_records = process_data(dfA, dfB)

                st.session_state.match_result_df = result_df.copy()
                st.session_state.manual_fill_records = manual_fill_records
                st.session_state.manual_selections = {}

                status_text.text("处理完成！")
                progress_bar.progress(100)

                # 显示统计信息
                total_count = len(result_df)
                matched_count = len(result_df[result_df["专业组代码"].notna() & (result_df["专业组代码"] != "")])
                manual_count = len(manual_fill_records)
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("总记录数", total_count)
                with col2:
                    st.metric("自动匹配成功", matched_count)
                with col3:
                    st.metric("需要手动补充", manual_count, delta=f"{manual_count}条")

                if manual_count > 0:
                    st.warning(f"⚠️ 发现 {manual_count} 条记录需要手动补充专业组代码")

            except Exception as e:
                st.error(f"处理错误：{e}")
                import traceback
                st.error(traceback.format_exc())

        # 显示手动补充界面（弹框形式）
        if st.session_state.match_result_df is not None and len(st.session_state.manual_fill_records) > 0:
            st.markdown("---")
            st.subheader("📝 手动补充专业组代码")
            
            # 省份筛选功能
            all_provinces = sorted(set([r.get("省份", "") for r in st.session_state.manual_fill_records if r.get("省份", "")]))
            all_provinces = [p for p in all_provinces if p]  # 过滤空值
            
            # 初始化省份筛选
            if 'selected_province' not in st.session_state:
                st.session_state.selected_province = "全部"
            
            # 省份筛选框
            col1, col2 = st.columns([1, 3])
            with col1:
                selected_province = st.selectbox(
                    "筛选省份",
                    ["全部"] + all_provinces,
                    index=0 if st.session_state.selected_province == "全部" else (all_provinces.index(st.session_state.selected_province) + 1 if st.session_state.selected_province in all_provinces else 0),
                    key="province_filter"
                )
                # 如果省份筛选改变，重置当前索引
                if selected_province != st.session_state.selected_province:
                    st.session_state.current_record_idx = 0
                st.session_state.selected_province = selected_province
            
            # 根据省份筛选记录（确保保留所有字段，包括候选记录）
            if selected_province == "全部":
                filtered_records = st.session_state.manual_fill_records
            else:
                # 使用列表推导式筛选，确保保留所有字段
                filtered_records = []
                for r in st.session_state.manual_fill_records:
                    if r.get("省份", "") == selected_province:
                        # 确保保留完整的记录，包括候选记录字段
                        filtered_records.append(r)
            
            # 显示筛选后的统计信息
            with col2:
                st.info(f"**筛选结果：** 共 {len(filtered_records)} 条记录需要手动补充（总记录数：{len(st.session_state.manual_fill_records)}）")
            
            if len(filtered_records) == 0:
                st.warning(f"⚠️ 省份「{selected_province}」没有需要手动补充的记录")
                st.stop()
            
            # 初始化当前处理的记录索引（基于筛选后的记录）
            if 'current_record_idx' not in st.session_state:
                st.session_state.current_record_idx = 0
            
            # 如果当前索引超出筛选后的记录范围，重置为0
            if st.session_state.current_record_idx >= len(filtered_records):
                st.session_state.current_record_idx = 0
            
            total_records = len(filtered_records)
            current_record = filtered_records[st.session_state.current_record_idx]
            idx = current_record["索引"]
            key = f"manual_select_{idx}"
            
            # 确保从原始记录中获取完整的候选记录信息
            # 如果筛选后的记录中候选记录丢失或为空，从原始记录中获取
            candidate_records_from_filtered = current_record.get("候选记录")
            if candidate_records_from_filtered is None or (isinstance(candidate_records_from_filtered, list) and len(candidate_records_from_filtered) == 0):
                # 从原始记录中查找对应的记录
                original_record = next((r for r in st.session_state.manual_fill_records if r.get("索引") == idx), None)
                if original_record:
                    original_candidates = original_record.get("候选记录")
                    if original_candidates is not None:
                        current_record["候选记录"] = original_candidates
                    else:
                        current_record["候选记录"] = []
                else:
                    current_record["候选记录"] = []
            
            # 显示进度
            if selected_province == "全部":
                progress_text = f"处理进度：{st.session_state.current_record_idx + 1} / {total_records}"
            else:
                progress_text = f"处理进度：{st.session_state.current_record_idx + 1} / {total_records}（省份：{selected_province}）"
            st.progress((st.session_state.current_record_idx + 1) / total_records, text=progress_text)
            
            # 弹框形式显示当前记录
            with st.expander(f"📋 记录 {st.session_state.current_record_idx + 1}：{current_record['学校名称']} - {current_record['招生专业']}", expanded=True):
                st.markdown("### 当前记录信息（专业分文件）")
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**学校名称：** {current_record['学校名称']}")
                    st.write(f"**省份：** {current_record['省份']}")
                    st.write(f"**招生专业：** {current_record['招生专业']}")
                    st.write(f"**一级层次：** {current_record['一级层次']}")
                with col2:
                    st.write(f"**招生科类：** {current_record['招生科类']}")
                    st.write(f"**招生批次：** {current_record['招生批次']}")
                    st.write(f"**招生类型：** {current_record['招生类型（选填）']}")
                    # 显示当前已选择的值（如果有）
                    current_value = st.session_state.manual_selections.get(key, "")
                    if current_value:
                        st.success(f"**已选择：** {current_value}")
                
                # 显示专业备注（选填）字段
                if current_record.get("专业备注（选填）"):
                    st.markdown("**专业备注（选填）：**")
                    st.info(current_record.get("专业备注（选填）", ""))
                
                st.markdown("---")
                st.markdown("### 招生计划中的候选记录")
                
                # 显示候选记录
                candidate_records = current_record.get("候选记录")
                # 处理None、空列表等情况
                if candidate_records is None:
                    candidate_records = []
                
                if candidate_records and len(candidate_records) > 0:
                    # 显示候选记录的详细信息表格
                    st.markdown("**候选记录详情：**")
                    candidate_df = pd.DataFrame(candidate_records)
                    # 重新排列列的顺序，专业组代码放在最前面
                    if '专业组代码' in candidate_df.columns:
                        cols = ['专业组代码'] + [c for c in candidate_df.columns if c != '专业组代码']
                        candidate_df = candidate_df[cols]
                    st.dataframe(candidate_df, use_container_width=True, hide_index=True)
                    
                    # 构建选项列表（显示专业组代码）
                    candidate_options = []
                    for i, cand in enumerate(candidate_records):
                        code = cand.get("专业组代码", "")
                        if code and str(code).strip():
                            candidate_options.append(str(code).strip())
                    
                    # 去重
                    candidate_options = list(set(candidate_options))
                    
                    if candidate_options:
                        # 添加"请选择"选项
                        options = ["请选择"] + candidate_options
                        # 获取当前选择（如果有）
                        current_selection = st.session_state.manual_selections.get(key, "请选择")
                        default_index = 0
                        if current_selection in options:
                            default_index = options.index(current_selection)
                        
                        selected_code = st.selectbox(
                            "选择专业组代码",
                            options,
                            index=default_index,
                            key=key
                        )
                        
                        if selected_code != "请选择":
                            st.session_state.manual_selections[key] = selected_code
                        else:
                            # 如果用户选择了"请选择"，清除之前的选择
                            if key in st.session_state.manual_selections:
                                del st.session_state.manual_selections[key]
                    else:
                        st.warning("⚠️ 候选记录中没有专业组代码，请手动输入")
                        input_key = f"{key}_input"
                        prev_value = st.session_state.get(input_key, "")
                        manual_input = st.text_input(
                            "手动输入专业组代码",
                            value=prev_value,
                            key=input_key
                        )
                        if manual_input and manual_input.strip():
                            st.session_state.manual_selections[key] = manual_input.strip()
                        elif key in st.session_state.manual_selections:
                            del st.session_state.manual_selections[key]
                else:
                    st.warning("⚠️ 该记录没有候选记录，请手动输入")
                    input_key = f"{key}_input"
                    prev_value = st.session_state.get(input_key, "")
                    manual_input = st.text_input(
                        "手动输入专业组代码",
                        value=prev_value,
                        key=input_key
                    )
                    if manual_input and manual_input.strip():
                        st.session_state.manual_selections[key] = manual_input.strip()
                    elif key in st.session_state.manual_selections:
                        del st.session_state.manual_selections[key]
            
            # 导航按钮
            col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
            with col1:
                if st.button("⏮️ 第一条", disabled=st.session_state.current_record_idx == 0):
                    st.session_state.current_record_idx = 0
                    st.rerun()
            with col2:
                if st.button("◀️ 上一条", disabled=st.session_state.current_record_idx == 0):
                    st.session_state.current_record_idx -= 1
                    st.rerun()
            with col3:
                if st.button("▶️ 下一条", disabled=st.session_state.current_record_idx >= total_records - 1):
                    st.session_state.current_record_idx += 1
                    st.rerun()
            with col4:
                if st.button("⏭️ 最后一条", disabled=st.session_state.current_record_idx >= total_records - 1):
                    st.session_state.current_record_idx = total_records - 1
                    st.rerun()
            
            st.markdown("---")
            
            # 应用所有手动选择并完成
            col1, col2 = st.columns([1, 1])
            with col1:
                if st.button("✅ 应用当前选择并继续", type="primary", use_container_width=True):
                    # 应用当前记录的选择
                    selected_code = None
                    if key in st.session_state.manual_selections:
                        selected_code = st.session_state.manual_selections[key]
                    elif f"{key}_input" in st.session_state:
                        input_value = st.session_state[f"{key}_input"]
                        if input_value and input_value.strip():
                            selected_code = input_value.strip()
                    
                    if selected_code and selected_code.strip():
                        updated_df = st.session_state.match_result_df.copy()
                        updated_df.at[idx, "专业组代码"] = selected_code.strip()
                        st.session_state.match_result_df = updated_df
                        st.success(f"✅ 已应用记录 {st.session_state.current_record_idx + 1} 的选择：{selected_code.strip()}")
                    
                    # 移动到下一条
                    if st.session_state.current_record_idx < total_records - 1:
                        st.session_state.current_record_idx += 1
                    st.rerun()
            
            with col2:
                if st.button("✅ 应用所有选择并完成", type="primary", use_container_width=True):
                    # 更新结果数据框
                    updated_df = st.session_state.match_result_df.copy()
                    applied_count = 0
                    
                    for record in st.session_state.manual_fill_records:
                        idx = record["索引"]
                        key = f"manual_select_{idx}"
                        input_key = f"{key}_input"
                        
                        # 检查是否有选择
                        selected_code = None
                        
                        # 先检查selectbox的选择
                        if key in st.session_state.manual_selections:
                            selected_code = st.session_state.manual_selections[key]
                            if selected_code == "请选择":
                                selected_code = None
                        elif key in st.session_state:
                            selected_code = st.session_state[key]
                            if selected_code == "请选择":
                                selected_code = None
                        
                        # 如果没有selectbox选择，检查text_input
                        if not selected_code and input_key in st.session_state:
                            input_value = st.session_state[input_key]
                            if input_value and input_value.strip():
                                selected_code = input_value.strip()
                        
                        # 应用选择
                        if selected_code and selected_code.strip():
                            updated_df.at[idx, "专业组代码"] = selected_code.strip()
                            applied_count += 1

                    st.session_state.match_result_df = updated_df
                    if applied_count > 0:
                        st.success(f"✅ 已应用 {applied_count} 条记录的手动选择！")
                    else:
                        st.warning("⚠️ 没有应用任何选择")
                    st.rerun()

        # 导出结果
        if st.session_state.match_result_df is not None:
            st.markdown("---")
            st.subheader("📥 导出结果")
            
            # 移除临时列
            export_df = st.session_state.match_result_df.drop(columns=["组合键"], errors='ignore')
            
            # 获取标题和年份
            headers = st.session_state.fileA_headers if st.session_state.fileA_headers else list(export_df.columns)
            year_value = st.session_state.fileB_year if st.session_state.fileB_year else ''
            
            # 导出结果到临时文件
            temp_output_path = "temp_match_result.xlsx"
            try:
                export_match_result_to_excel(export_df, headers, year_value, temp_output_path)
                
                # 读取文件并转换为base64
                with open(temp_output_path, "rb") as f:
                    bytes_data = f.read()
                b64 = base64.b64encode(bytes_data).decode()
                href = f'<a href="data:application/octet-stream;base64,{b64}" download="专业组代码匹配结果.xlsx">点击下载匹配结果</a>'
                st.markdown(href, unsafe_allow_html=True)
                
                # 清理临时文件
                if os.path.exists(temp_output_path):
                    os.remove(temp_output_path)
            except Exception as e:
                st.error(f"导出失败：{str(e)}")
                import traceback
                st.error(traceback.format_exc())

            # 清理临时文件按钮
            if st.button("清理临时文件", key="cleanup_temp"):
                if st.session_state.temp_fileA_path and os.path.exists(st.session_state.temp_fileA_path):
                    os.remove(st.session_state.temp_fileA_path)
                if st.session_state.temp_fileB_path and os.path.exists(st.session_state.temp_fileB_path):
                    os.remove(st.session_state.temp_fileB_path)
                st.session_state.temp_fileA_path = None
                st.session_state.temp_fileB_path = None
                st.success("临时文件已清理")

    else:
        st.info("请先上传两个Excel文件")

# ====================== tab5：网页图片提取PDF ======================
with tab6:
    st.header("就业质量报告图片提取")

    url = st.text_input("请输入就业质量报告网页链接", placeholder="例如：https://www.example.com/report.html")

    if st.button("开始提取图片"):
        if not url:
            st.warning("请输入有效的网页链接")
        else:
            output_folder = tempfile.mkdtemp()
            with st.spinner("正在抓取图片..."):
                try:
                    image_paths = fetch_images_static(url, output_folder)
                except Exception as e:
                    st.error(f"抓取失败: {e}")
                    image_paths = []

            if image_paths:
                st.success(f"成功提取到 {len(image_paths)} 张图片")

                with st.expander(f"点击查看 {len(image_paths)} 张图片预览", expanded=False):
                    cols = st.columns(5)
                    for i, path in enumerate(image_paths):
                        cols[i % 5].image(path, width=120)

                pdf_path = os.path.join(output_folder, "图片合集.pdf")
                if images_to_pdf(image_paths, pdf_path):
                    with open(pdf_path, "rb") as f:
                        st.download_button("📥 下载合成PDF", f, file_name="就业质量报告.pdf", mime="application/pdf")
                else:
                    st.warning("PDF合成失败")
            else:
                st.warning("未抓取到任何图片")





# ====================== tab7：招生计划工具======================
with tab7:
    st.header("招生计划数据比对与转换工具")
    st.markdown("上传招生计划、专业分和院校分文件进行比对，导出未匹配数据为专业分/院校分格式")

    # 初始化session state
    if 'plan_data' not in st.session_state:
        st.session_state.plan_data = None
    if 'score_data' not in st.session_state:
        st.session_state.score_data = None
    if 'college_data' not in st.session_state:
        st.session_state.college_data = None
    if 'plan_score_results' not in st.session_state:
        st.session_state.plan_score_results = []
    if 'plan_college_results' not in st.session_state:
        st.session_state.plan_college_results = []

    # 工作流步骤显示
    col1, col2, col3, col4, col5 = st.columns([1, 0.3, 1, 0.3, 1])
    with col1:
        st.markdown("""
        <div style="text-align: center; padding: 10px; background: #e3f2fd; border-radius: 10px;">
            <div style="font-size: 24px; font-weight: bold;">1</div>
            <div>上传文件</div>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div style="text-align: center; padding: 10px; background: #f5f5f5; border-radius: 10px;">
            <div style="font-size: 24px; font-weight: bold;">2</div>
            <div>数据比对</div>
        </div>
        """, unsafe_allow_html=True)
    with col5:
        st.markdown("""
        <div style="text-align: center; padding: 10px; background: #f5f5f5; border-radius: 10px;">
            <div style="font-size: 24px; font-weight: bold;">3</div>
            <div>导出未匹配数据</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # 字段说明
    with st.expander("📋 比对字段说明", expanded=False):
        st.markdown("""
        **比对1（招生计划 vs 专业分）：** 检查招生计划的记录是否在专业分中存在
        - 匹配字段：年份、省份、学校、科类、批次、专业、层次、专业组代码

        **比对2（招生计划 vs 院校分）：** 检查招生计划的记录是否在院校分中存在
        - 匹配字段：年份、省份、学校、科类、批次、专业组代码
        """)

    # 文件上传区域
    col1, col2, col3 = st.columns(3)

    with col1:
        st.subheader("招生计划文件")
        plan_file = st.file_uploader("上传招生计划文件", type=["xlsx", "xls"], key="tab7_plan_file")
        if plan_file is not None:
            try:
                plan_df = pd.read_excel(plan_file, engine='openpyxl')
                st.session_state.plan_data = plan_df
                st.success(f"✓ 文件加载成功\n文件名: {plan_file.name}\n记录数: {len(plan_df)} 条")
            except Exception as e:
                st.error(f"❌ 文件读取失败: {str(e)}")

    with col2:
        st.subheader("专业分文件")
        score_file = st.file_uploader("上传专业分文件", type=["xlsx", "xls"], key="tab7_score_file")
        if score_file is not None:
            try:
                score_df = pd.read_excel(score_file, engine='openpyxl')
                st.session_state.score_data = score_df
                st.success(f"✓ 文件加载成功\n文件名: {score_file.name}\n记录数: {len(score_df)} 条")
            except Exception as e:
                st.error(f"❌ 文件读取失败: {str(e)}")

    with col3:
        st.subheader("院校分文件")
        college_file = st.file_uploader("上传院校分文件", type=["xlsx", "xls"], key="tab7_college_file")
        if college_file is not None:
            try:
                college_df = pd.read_excel(college_file, engine='openpyxl')
                st.session_state.college_data = college_df
                st.success(f"✓ 文件加载成功\n文件名: {college_file.name}\n记录数: {len(college_df)} 条")
            except Exception as e:
                st.error(f"❌ 文件读取失败: {str(e)}")

    st.markdown("---")

    # 比对按钮
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    with col1:
        compare_plan_score_btn = st.button("比对1：招生计划 vs 专业分", type="primary", use_container_width=True)
    with col2:
        compare_plan_college_btn = st.button("比对2：招生计划 vs 院校分", type="primary", use_container_width=True)
    with col3:
        compare_all_btn = st.button("全部比对", type="primary", use_container_width=True)
    with col4:
        reset_btn = st.button("重置", use_container_width=True)

    # 执行比对
    if compare_plan_score_btn:
        if st.session_state.plan_data is None:
            st.error("请先上传招生计划文件")
        elif st.session_state.score_data is None:
            st.error("请先上传专业分文件")
        else:
            with st.spinner("正在比对数据..."):
                st.session_state.plan_score_results = compare_plan_vs_score(
                    st.session_state.plan_data, st.session_state.score_data
                )
            st.success("比对1完成！")
            st.balloons()

    if compare_plan_college_btn:
        if st.session_state.plan_data is None:
            st.error("请先上传招生计划文件")
        elif st.session_state.college_data is None:
            st.error("请先上传院校分文件")
        else:
            with st.spinner("正在比对数据..."):
                st.session_state.plan_college_results = compare_plan_vs_college(
                    st.session_state.plan_data, st.session_state.college_data
                )
            st.success("比对2完成！")
            st.balloons()

    if compare_all_btn:
        comparisons = []
        if st.session_state.plan_data is not None and st.session_state.score_data is not None:
            comparisons.append("比对1")
        if st.session_state.plan_data is not None and st.session_state.college_data is not None:
            comparisons.append("比对2")

        if len(comparisons) == 0:
            st.error("请至少上传两个文件以进行比对")
        else:
            with st.spinner("正在执行全部比对..."):
                if "比对1" in comparisons:
                    st.session_state.plan_score_results = compare_plan_vs_score(
                        st.session_state.plan_data, st.session_state.score_data
                    )
                if "比对2" in comparisons:
                    st.session_state.plan_college_results = compare_plan_vs_college(
                        st.session_state.plan_data, st.session_state.college_data
                    )
            st.success("全部比对完成！")
            st.balloons()

    if reset_btn:
        st.session_state.plan_data = None
        st.session_state.score_data = None
        st.session_state.college_data = None
        st.session_state.plan_score_results = []
        st.session_state.plan_college_results = []
        st.success("重置完成！")
        st.rerun()

    # 显示比对结果
    if len(st.session_state.plan_score_results) > 0 or len(st.session_state.plan_college_results) > 0:
        st.markdown("---")

        # 创建标签页
        tab_plan_score, tab_plan_college = st.tabs([
            "比对1：招生计划 vs 专业分",
            "比对2：招生计划 vs 院校分"
        ])

        # 比对1结果
        with tab_plan_score:
            if len(st.session_state.plan_score_results) > 0:
                results = st.session_state.plan_score_results
                total = len(results)
                matched = sum(1 for r in results if r['exists'])
                unmatched = total - matched
                rate = (matched / total * 100) if total > 0 else 0

                # 统计信息
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("总记录数", total)
                with col2:
                    st.metric("匹配记录数", matched, delta=f"{rate:.1f}%")
                with col3:
                    st.metric("未匹配记录数", unmatched)
                with col4:
                    st.metric("匹配率", f"{rate:.1f}%")

                # 筛选控件
                st.markdown("### 筛选条件")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    provinces = sorted(set(r['keyFields']['省份'] for r in results if r['keyFields']['省份']))
                    province_filter = st.selectbox("省份", ["全部"] + provinces, key="ps_province")
                with col2:
                    batches = sorted(set(r['keyFields']['批次'] for r in results if r['keyFields']['批次']))
                    batch_filter = st.selectbox("批次", ["全部"] + batches, key="ps_batch")
                with col3:
                    match_status_filter = st.selectbox("匹配状态", ["全部", "匹配", "未匹配"], key="ps_status")
                with col4:
                    display_option = st.selectbox("显示选项", ["全部", "前100条", "前500条"], key="ps_display")

                # 应用筛选
                filtered_results = results
                if province_filter != "全部":
                    filtered_results = [r for r in filtered_results if r['keyFields']['省份'] == province_filter]
                if batch_filter != "全部":
                    filtered_results = [r for r in filtered_results if r['keyFields']['批次'] == batch_filter]
                if match_status_filter == "匹配":
                    filtered_results = [r for r in filtered_results if r['exists']]
                elif match_status_filter == "未匹配":
                    filtered_results = [r for r in filtered_results if not r['exists']]

                display_count = len(filtered_results)
                if display_option == "前100条":
                    display_count = min(100, len(filtered_results))
                elif display_option == "前500条":
                    display_count = min(500, len(filtered_results))

                # 显示表格
                st.markdown(
                    f"### 比对结果（显示 {min(display_count, len(filtered_results))} / {len(filtered_results)} 条）")
                display_results = filtered_results[:display_count]

                if len(display_results) > 0:
                    # 准备表格数据
                    table_data = []
                    for r in display_results:
                        table_data.append({
                            '序号': r['index'],
                            '年份': r['keyFields']['年份'],
                            '省份': r['keyFields']['省份'],
                            '学校': r['keyFields']['学校'],
                            '科类': r['keyFields']['科类'],
                            '批次': r['keyFields']['批次'],
                            '专业': r['keyFields']['专业'],
                            '层次': r['keyFields']['层次'],
                            '专业组代码': r['keyFields']['专业组代码'] or '-',
                            '招生人数': r['otherInfo']['招生人数'] or '-',
                            '匹配状态': '✓ 存在' if r['exists'] else '✗ 不存在'
                        })

                    df_display = pd.DataFrame(table_data)
                    st.dataframe(df_display, use_container_width=True, hide_index=True)

                # 导出按钮
                if st.button("导出比对1结果", key="export_ps", use_container_width=True):
                    try:
                        export_data = []
                        for r in results:
                            export_data.append({
                                '序号': r['index'],
                                '年份': r['keyFields']['年份'],
                                '省份': r['keyFields']['省份'],
                                '学校': r['keyFields']['学校'],
                                '科类': r['keyFields']['科类'],
                                '批次': r['keyFields']['批次'],
                                '专业': r['keyFields']['专业'],
                                '层次': r['keyFields']['层次'],
                                '专业组代码': r['keyFields']['专业组代码'],
                                '招生人数': r['otherInfo']['招生人数'],
                                '学费': r['otherInfo']['学费'],
                                '学制': r['otherInfo']['学制'],
                                '专业代码': r['otherInfo']['专业代码'],
                                '匹配状态': '存在' if r['exists'] else '不存在',
                                '匹配说明': '该记录在专业分文件中存在' if r['exists'] else '该记录在专业分文件中不存在'
                            })

                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            pd.DataFrame(export_data).to_excel(writer, index=False, sheet_name='比对1_招生计划vs专业分')

                        output.seek(0)
                        st.download_button(
                            "📥 下载比对1结果",
                            output,
                            file_name=f"比对1_招生计划vs专业分_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"导出失败: {str(e)}")
            else:
                st.info("暂无比对结果，请先执行比对")

        # 比对2结果
        with tab_plan_college:
            if len(st.session_state.plan_college_results) > 0:
                results = st.session_state.plan_college_results
                total = len(results)
                matched = sum(1 for r in results if r['exists'])
                unmatched = total - matched
                rate = (matched / total * 100) if total > 0 else 0

                # 统计信息
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("总记录数", total)
                with col2:
                    st.metric("匹配记录数", matched, delta=f"{rate:.1f}%")
                with col3:
                    st.metric("未匹配记录数", unmatched)
                with col4:
                    st.metric("匹配率", f"{rate:.1f}%")

                # 筛选控件
                st.markdown("### 筛选条件")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    provinces = sorted(set(r['keyFields']['省份'] for r in results if r['keyFields']['省份']))
                    province_filter = st.selectbox("省份", ["全部"] + provinces, key="pc_province")
                with col2:
                    batches = sorted(set(r['keyFields']['批次'] for r in results if r['keyFields']['批次']))
                    batch_filter = st.selectbox("批次", ["全部"] + batches, key="pc_batch")
                with col3:
                    match_status_filter = st.selectbox("匹配状态", ["全部", "匹配", "未匹配"], key="pc_status")
                with col4:
                    display_option = st.selectbox("显示选项", ["全部", "前100条", "前500条"], key="pc_display")

                # 应用筛选
                filtered_results = results
                if province_filter != "全部":
                    filtered_results = [r for r in filtered_results if r['keyFields']['省份'] == province_filter]
                if batch_filter != "全部":
                    filtered_results = [r for r in filtered_results if r['keyFields']['批次'] == batch_filter]
                if match_status_filter == "匹配":
                    filtered_results = [r for r in filtered_results if r['exists']]
                elif match_status_filter == "未匹配":
                    filtered_results = [r for r in filtered_results if not r['exists']]

                display_count = len(filtered_results)
                if display_option == "前100条":
                    display_count = min(100, len(filtered_results))
                elif display_option == "前500条":
                    display_count = min(500, len(filtered_results))

                # 显示表格
                st.markdown(
                    f"### 比对结果（显示 {min(display_count, len(filtered_results))} / {len(filtered_results)} 条）")
                display_results = filtered_results[:display_count]

                if len(display_results) > 0:
                    # 准备表格数据
                    table_data = []
                    for r in display_results:
                        table_data.append({
                            '序号': r['index'],
                            '省份': r['keyFields']['省份'],
                            '学校': r['keyFields']['学校'],
                            '科类': r['keyFields']['科类'],
                            '批次': r['keyFields']['批次'],
                            '专业组代码': r['keyFields']['专业组代码'] or '-',
                            '招生代码': r['keyFields']['招生代码'] or '-',
                            '专业': r['otherInfo']['专业'] or '-',
                            '匹配状态': '✓ 存在' if r['exists'] else '✗ 不存在'
                        })

                    df_display = pd.DataFrame(table_data)
                    st.dataframe(df_display, use_container_width=True, hide_index=True)

                # 导出按钮
                if st.button("导出比对2结果", key="export_pc", use_container_width=True):
                    try:
                        export_data = []
                        for r in results:
                            export_data.append({
                                '序号': r['index'],
                                '年份': r['otherInfo']['年份'],
                                '省份': r['keyFields']['省份'],
                                '学校': r['keyFields']['学校'],
                                '科类': r['keyFields']['科类'],
                                '批次': r['keyFields']['批次'],
                                '专业组代码': r['keyFields']['专业组代码'],
                                '招生代码': r['keyFields']['招生代码'],
                                '专业': r['otherInfo']['专业'],
                                '层次': r['otherInfo']['层次'],
                                '招生人数': r['otherInfo']['招生人数'],
                                '匹配状态': '存在' if r['exists'] else '不存在',
                                '匹配说明': '该记录在院校分文件中存在' if r['exists'] else '该记录在院校分文件中不存在'
                            })

                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            pd.DataFrame(export_data).to_excel(writer, index=False, sheet_name='比对2_招生计划vs院校分')

                        output.seek(0)
                        st.download_button(
                            "📥 下载比对2结果",
                            output,
                            file_name=f"比对2_招生计划vs院校分_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"导出失败: {str(e)}")
            else:
                st.info("暂无比对结果，请先执行比对")

        # 全局导出区域
        if len(st.session_state.plan_score_results) > 0 or len(st.session_state.plan_college_results) > 0:
            st.markdown("---")
            st.markdown("### 📤 全局导出功能")

            # 收集所有未匹配的数据
            all_unmatched_results = []
            if len(st.session_state.plan_score_results) > 0:
                all_unmatched_results.extend([r for r in st.session_state.plan_score_results if not r['exists']])
            if len(st.session_state.plan_college_results) > 0:
                all_unmatched_results.extend([r for r in st.session_state.plan_college_results if not r['exists']])

            # 使用三列布局，添加院校分格式导出
            col1, col2, col3 = st.columns([1, 1, 1])

            with col1:
                if st.button("📊 导出全部结果", use_container_width=True):
                    try:
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            # 比对1结果
                            if len(st.session_state.plan_score_results) > 0:
                                export_data = []
                                for r in st.session_state.plan_score_results:
                                    export_data.append({
                                        '序号': r['index'],
                                        '年份': r['keyFields']['年份'],
                                        '省份': r['keyFields']['省份'],
                                        '学校': r['keyFields']['学校'],
                                        '科类': r['keyFields']['科类'],
                                        '批次': r['keyFields']['批次'],
                                        '专业': r['keyFields']['专业'],
                                        '层次': r['keyFields']['层次'],
                                        '专业组代码': r['keyFields']['专业组代码'],
                                        '招生人数': r['otherInfo']['招生人数'],
                                        '学费': r['otherInfo']['学费'],
                                        '学制': r['otherInfo']['学制'],
                                        '专业代码': r['otherInfo']['专业代码'],
                                        '匹配状态': '存在' if r['exists'] else '不存在',
                                        '匹配说明': '该记录在专业分文件中存在' if r['exists'] else '该记录在专业分文件中不存在'
                                    })
                                pd.DataFrame(export_data).to_excel(writer, index=False,
                                                                   sheet_name='比对1_招生计划vs专业分')

                            # 比对2结果
                            if len(st.session_state.plan_college_results) > 0:
                                export_data = []
                                for r in st.session_state.plan_college_results:
                                    export_data.append({
                                        '序号': r['index'],
                                        '年份': r['otherInfo']['年份'],
                                        '省份': r['keyFields']['省份'],
                                        '学校': r['keyFields']['学校'],
                                        '科类': r['keyFields']['科类'],
                                        '批次': r['keyFields']['批次'],
                                        '专业组代码': r['keyFields']['专业组代码'],
                                        '招生代码': r['keyFields']['招生代码'],
                                        '专业': r['otherInfo']['专业'],
                                        '层次': r['otherInfo']['层次'],
                                        '招生人数': r['otherInfo']['招生人数'],
                                        '匹配状态': '存在' if r['exists'] else '不存在',
                                        '匹配说明': '该记录在院校分文件中存在' if r['exists'] else '该记录在院校分文件中不存在'
                                    })
                                pd.DataFrame(export_data).to_excel(writer, index=False,
                                                                   sheet_name='比对2_招生计划vs院校分')

                            # 统计报告
                            summary_data = {
                                '比对类型': ['比对1：招生计划 vs 专业分', '比对2：招生计划 vs 院校分'],
                                '总记录数': [
                                    len(st.session_state.plan_score_results),
                                    len(st.session_state.plan_college_results)
                                ],
                                '匹配记录数': [
                                    sum(1 for r in st.session_state.plan_score_results if r['exists']),
                                    sum(1 for r in st.session_state.plan_college_results if r['exists'])
                                ],
                                '匹配率': [
                                    f"{(sum(1 for r in st.session_state.plan_score_results if r['exists']) / len(st.session_state.plan_score_results) * 100):.1f}%" if len(
                                        st.session_state.plan_score_results) > 0 else "0%",
                                    f"{(sum(1 for r in st.session_state.plan_college_results if r['exists']) / len(st.session_state.plan_college_results) * 100):.1f}%" if len(
                                        st.session_state.plan_college_results) > 0 else "0%"
                                ]
                            }
                            pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name='统计报告')

                        output.seek(0)
                        st.download_button(
                            "📥 下载全部结果",
                            output,
                            file_name=f"数据比对结果汇总_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"导出失败: {str(e)}")

            with col2:
                if len(all_unmatched_results) > 0:
                    if st.button("⭐ 导出未匹配数据为专业分格式", type="primary", use_container_width=True):
                        try:
                            # 提取原始数据（去重，因为同一个记录可能在比对1和比对2中都未匹配）
                            seen_indices = set()
                            conversion_data = []
                            for r in all_unmatched_results:
                                original_idx = r['originalIndex']
                                if original_idx not in seen_indices:
                                    seen_indices.add(original_idx)
                                    conversion_data.append(st.session_state.plan_data.iloc[original_idx].to_dict())

                            # 转换数据
                            converted_data = convert_data(conversion_data)

                            # 导出
                            output = BytesIO()
                            temp_path = "temp_converted.xlsx"
                            export_converted_data_to_excel(converted_data, conversion_data, temp_path)

                            with open(temp_path, 'rb') as f:
                                st.download_button(
                                    "📥 下载转换后的专业分数据",
                                    f.read(),
                                    file_name=f"专业分数据_未匹配数据_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                            os.remove(temp_path)
                            st.success(f"转换完成！共转换 {len(converted_data)} 条数据（已去重）")
                        except Exception as e:
                            st.error(f"转换失败: {str(e)}")
                else:
                    st.info("暂无未匹配数据")

            with col3:
                if len(all_unmatched_results) > 0:
                    if st.button("⭐ 导出未匹配数据为院校分格式", type="primary", use_container_width=True):
                        try:
                            # 检查是否有院校分数据
                            if 'college_data' not in st.session_state or st.session_state.college_data is None:
                                st.error("请先上传院校分文件，以便进行比对过滤")
                            else:
                                # 提取原始招生计划数据
                                plan_df = st.session_state.plan_data.copy()
                                college_df = st.session_state.college_data.copy()
                                
                                # 使用新的过滤函数，只导出招生计划中不存在于院校分的数据
                                unmatched_records = filter_unmatched_plan_data_for_college_export(plan_df, college_df)
                                
                                if len(unmatched_records) == 0:
                                    st.warning("⚠️ 所有招生计划数据都已存在于院校分中，无需转换")
                                else:
                                    # 提取未匹配数据
                                    conversion_data = [r['data'] for r in unmatched_records]
                                    
                                    # 转换数据为院校分格式
                                    college_score_data = convert_to_college_score_format(conversion_data)
                                    
                                    # 导出
                                    temp_path = "temp_college_score.xlsx"
                                    export_college_score_data_to_excel(college_score_data, conversion_data, temp_path)
                                    
                                    with open(temp_path, 'rb') as f:
                                        st.download_button(
                                            "📥 下载转换后的院校分数据",
                                            f.read(),
                                            file_name=f"院校分数据_未匹配数据导出{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                        )
                                    
                                    os.remove(temp_path)
                                    st.success(f"转换完成！共转换 {len(college_score_data)} 条数据")
                        except Exception as e:
                            st.error(f"转换失败: {str(e)}")
                            import traceback
                            st.error(traceback.format_exc())
                else:
                    st.info("暂无未匹配数据")

# 页脚
st.markdown("---")
st.markdown("© 数据处理", unsafe_allow_html=True)
